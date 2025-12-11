#!/usr/bin/env python3
# limpiador_excel_gui.py (versión corregida con soporte para fechas tipo "05-DEC-24")
import os
import sys
import subprocess
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from pathlib import Path
from datetime import datetime
import re
import copy
import shutil
import logging
import traceback
import gc

# Configuración de logging
LOG_DIR = "logs"
os.makedirs(LOG_DIR, exist_ok=True)
log_filename = os.path.join(LOG_DIR, f"limpiador_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Dependencias externas
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import numbers, Alignment
    logger.info("openpyxl cargado correctamente")
except Exception as e:
    logger.critical(f"Error al importar openpyxl: {e}")
    raise ImportError("Instala openpyxl: pip install openpyxl") from e

try:
    from dateutil import parser as dateparser
    logger.info("python-dateutil cargado correctamente")
except Exception as e:
    logger.critical(f"Error al importar python-dateutil: {e}")
    raise ImportError("Instala python-dateutil: pip install python-dateutil") from e

# xlwings (opcional)
USE_XLWINGS = False
try:
    import xlwings as xw
    USE_XLWINGS = True
    logger.info("xlwings disponible")
except Exception as e:
    USE_XLWINGS = False
    logger.warning(f"xlwings no disponible: {e}")

# Carpetas por defecto
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CARPETA_LIMPIOS = "limpios"
os.makedirs(CARPETA_LIMPIOS, exist_ok=True)

# Configuración para archivos grandes
MAX_FILE_SIZE_MB = 100  # Tamaño máximo recomendado sin advertencia
CHUNK_SIZE = 100  # Procesar hojas en chunks para archivos grandes

# Comportamiento para enteros largos (>15 dígitos)
# Si False: preservarlos como texto (seguro). Si True: forzar conversión a número (puede perder precisión).
FORCE_NUMERIC_LONGS = False

# ---------------------------
# Utilidades de parseo/limpieza
# ---------------------------
NBSP = '\xa0'

# Mapeo de meses en inglés a números
MONTH_MAP = {
    'JAN': '01', 'FEB': '02', 'MAR': '03', 'APR': '04', 'MAY': '05', 'JUN': '06',
    'JUL': '07', 'AUG': '08', 'SEP': '09', 'OCT': '10', 'NOV': '11', 'DEC': '12'
}

# Agregar esta línea para manejar tanto mayúsculas como minúsculas
MONTH_MAP.update({k.lower(): v for k, v in MONTH_MAP.items()})

def clean_text(s):
    """Limpia texto de manera segura"""
    try:
        if s is None:
            return ""
        if not isinstance(s, str):
            return s
        s2 = s.strip().replace(NBSP, "").replace('\t', "").replace('\r', "").replace('\n', "")
        return s2
    except Exception as e:
        logger.warning(f"Error limpiando texto: {e}")
        return ""

def try_parse_number(s):
    """Intenta parsear un número de manera robusta.

    Retorna (value, tag) donde tag puede ser:
      - 'General' -> valor numérico seguro (int/float)
      - 'AS_TEXT' -> devolver como texto para preservar precisión (enteros largos)
      - None -> no es número
    """
    try:
        if s is None:
            return None, None
        if isinstance(s, (int, float)):
            # Si ya es número asumimos 'General'
            return s, 'General'
        orig = str(s).strip()
        if orig == "":
            return None, None

        # eliminar caracteres invisibles y espacios
        s1 = orig.replace(NBSP, "").replace(' ', '')

        # quitar prefijo '#' si existe (pero conservar resto)
        if s1.startswith('#'):
            s1 = s1.lstrip('#')

        # normalizar separadores decimales
        cnt_dot = s1.count('.')
        cnt_comma = s1.count(',')
        s2 = s1
        if cnt_dot > 0 and cnt_comma > 0:
            s2 = s2.replace('.', '').replace(',', '.')
        elif cnt_dot > 1 and cnt_comma == 0:
            s2 = s2.replace('.', '')
        elif cnt_comma > 1 and cnt_dot == 0:
            s2 = s2.replace(',', '')
        elif cnt_dot == 1 and cnt_comma == 0:
            part_after = s2.split('.')[-1]
            if len(part_after) == 3:
                s2 = s2.replace('.', '')
        elif cnt_comma == 1 and cnt_dot == 0:
            part_after = s2.split(',')[-1]
            if len(part_after) == 3:
                s2 = s2.replace(',', '')
            else:
                s2 = s2.replace(',', '.')

        # conservar solo dígitos, punto y signo
        clean = ''.join(ch for ch in s2 if (ch.isdigit() or ch in '.-'))
        if clean == "" or clean in ['.', '-', '-.']:
            return None, None

        # Si contiene punto -> float (comprobar parte entera)
        if '.' in clean:
            int_part = clean.split('.')[0].lstrip('-')
            # Si la parte entera es mayor a 15 dígitos, devolver como texto para evitar pérdida de precisión
            if len(int_part) > 15:
                return clean, 'AS_TEXT'
            try:
                val = float(clean)
                return val, 'General'
            except Exception:
                return clean, 'AS_TEXT'

        # Entero puro: revisar longitud de dígitos
        digits = clean.lstrip('-')
        if len(digits) > 15:
            # demasiados dígitos: devolver como texto
            return digits, 'AS_TEXT'
        try:
            val = int(clean)
            return val, 'General'
        except Exception:
            return clean, 'AS_TEXT'
    except Exception as e:
        logger.debug(f"Error parseando número '{s}': {e}")
        return None, None

def parse_english_month_date(date_str):
    """Convierte fechas en formato como '05-DEC-24' a objeto datetime"""
    try:
        if not isinstance(date_str, str):
            return None
            
        # Patrón para detectar formato DD-MMM-AA o DD-MMM-AAAA
        pattern = r'^(\d{1,2})[-/\s\.]?([A-Za-z]{3})[-/\s\.]?(\d{2,4})$'
        match = re.match(pattern, date_str.upper().strip())
        
        if match:
            day = match.group(1)
            month_str = match.group(2).upper()  # Convertir a mayúsculas para el mapeo
            year = match.group(3)
            
            # Convertir mes a número - AHORA FUNCIONA CON MAYÚSCULAS Y MINÚSCULAS
            if month_str.upper() in MONTH_MAP:  # Cambiar esta línea
                month_num = MONTH_MAP[month_str.upper()]  # Cambiar esta línea
                
                # Ajustar año (si es 2 dígitos, asumir siglo 20/21)
                if len(year) == 2:
                    year_num = int(year)
                    # Si el año es menor a 50, asumir 2000+, si no 1900+
                    full_year = 2000 + year_num if year_num < 50 else 1900 + year_num
                else:
                    full_year = int(year)
                
                # Crear fecha
                try:
                    return datetime(full_year, int(month_num), int(day))
                except ValueError as e:
                    logger.debug(f"Fecha inválida {day}-{month_num}-{full_year}: {e}")
                    return None
                    
    except Exception as e:
        logger.debug(f"Error parseando fecha inglesa '{date_str}': {e}")
    
    return None

def try_parse_date(s):
    """Intenta parsear una fecha de manera segura"""
    try:
        if s is None:
            return None
        if isinstance(s, datetime):
            return s
        if isinstance(s, (int, float)):
            return None
        text = str(s).strip()
        if text == "":
            return None
        
        # Primero intentar con el formato de mes en inglés
        english_date = parse_english_month_date(text)
        if english_date:
            return english_date
            
        # Si no funciona, intentar con dateparser
        dt = dateparser.parse(text, dayfirst=True, yearfirst=False, fuzzy=False)
        return dt
    except Exception as e:
        logger.debug(f"Error parseando fecha '{s}': {e}")
        return None

# ---------------------------
# Copiar estilos y visuales (con manejo de errores mejorado)
# ---------------------------

DEFAULT_NUMBER_FORMATS = {
    'General', 'general', '@', '0', '0.00', '0%', '0.00%', '#,##0', '#,##0.00',
    'mm-dd-yy', 'm/d/yy', 'd/m/yy', 'yyyy-mm-dd', 'yyyy-mm-dd hh:mm:ss', 'dd/mm/yyyy'
}

def is_custom_number_format(fmt):
    """Verifica si un formato es personalizado"""
    try:
        if not fmt:
            return False
        fmt_str = str(fmt).strip()
        if '%' in fmt_str:
            return True
        return fmt_str not in DEFAULT_NUMBER_FORMATS
    except Exception as e:
        logger.debug(f"Error verificando formato: {e}")
        return False

def copy_cell_style(src_cell, tgt_cell):
    """Copia estilos de celda de manera segura"""
    try:
        if src_cell.font:
            tgt_cell.font = copy.copy(src_cell.font)
    except Exception as e:
        logger.debug(f"Error copiando fuente: {e}")
    
    try:
        if src_cell.fill:
            tgt_cell.fill = copy.copy(src_cell.fill)
    except Exception as e:
        logger.debug(f"Error copiando relleno: {e}")
    
    try:
        if src_cell.border:
            tgt_cell.border = copy.copy(src_cell.border)
    except Exception as e:
        logger.debug(f"Error copiando borde: {e}")
    
    try:
        if src_cell.alignment:
            tgt_cell.alignment = copy.copy(src_cell.alignment)
    except Exception as e:
        logger.debug(f"Error copiando alineación: {e}")
    
    try:
        if src_cell.protection:
            tgt_cell.protection = copy.copy(src_cell.protection)
    except Exception as e:
        logger.debug(f"Error copiando protección: {e}")

def copy_sheet_visuals(ws_src_styles, ws_src_values, ws_tgt):
    """Copia visuales de hoja con manejo robusto de errores"""
    try:
        # Column widths & hidden
        for col, dim in ws_src_styles.column_dimensions.items():
            try:
                tgt_dim = ws_tgt.column_dimensions[col]
                if dim.width is not None:
                    tgt_dim.width = dim.width
                tgt_dim.hidden = dim.hidden
            except Exception as e:
                logger.debug(f"Error copiando dimensión columna {col}: {e}")
    except Exception as e:
        logger.warning(f"Error procesando dimensiones de columna: {e}")

    try:
        # Row heights & hidden
        for r, dim in ws_src_styles.row_dimensions.items():
            try:
                tgt_dim = ws_tgt.row_dimensions[r]
                if dim.height is not None:
                    tgt_dim.height = dim.height
                tgt_dim.hidden = dim.hidden
            except Exception as e:
                logger.debug(f"Error copiando dimensión fila {r}: {e}")
    except Exception as e:
        logger.warning(f"Error procesando dimensiones de fila: {e}")

    try:
        # Freeze panes
        src_freeze = getattr(ws_src_styles, "freeze_panes", None)
        if src_freeze:
            ws_tgt.freeze_panes = src_freeze
    except Exception as e:
        logger.warning(f"Error copiando freeze_panes: {e}")

    try:
        # Merged cells
        for merge in ws_src_styles.merged_cells.ranges:
            try:
                ws_tgt.merge_cells(str(merge))
            except Exception as e:
                logger.debug(f"Error mergeando celda {merge}: {e}")
    except Exception as e:
        logger.warning(f"Error procesando celdas mergeadas: {e}")

    try:
        # Tab color
        if ws_src_styles.sheet_properties.tabColor is not None:
            ws_tgt.sheet_properties.tabColor = copy.copy(ws_src_styles.sheet_properties.tabColor)
    except Exception as e:
        logger.debug(f"Error copiando color de tab: {e}")

    # Copiar celdas con procesamiento optimizado
    try:
        max_r = min(ws_src_styles.max_row, 1048576)  # Límite Excel
        max_c = min(ws_src_styles.max_column, 16384)  # Límite Excel
        
        logger.info(f"Procesando hoja '{ws_src_styles.title}': {max_r} filas x {max_c} columnas")
        
        # Procesar en chunks para archivos grandes
        chunk_rows = 1000
        for start_row in range(1, max_r + 1, chunk_rows):
            end_row = min(start_row + chunk_rows - 1, max_r)
            
            for r in range(start_row, end_row + 1):
                for c in range(1, max_c + 1):
                    try:
                        process_cell(ws_src_styles, ws_src_values, ws_tgt, r, c)
                    except Exception as e:
                        logger.debug(f"Error procesando celda ({r},{c}): {e}")
                        # Continuar con la siguiente celda
            
            # Liberar memoria periódicamente
            if start_row % 5000 == 0:
                gc.collect()
                
    except Exception as e:
        logger.error(f"Error crítico copiando celdas: {e}")
        raise

def process_cell(ws_src_styles, ws_src_values, ws_tgt, r, c):
    """Procesa una celda individual de manera segura"""
    try:
        cell_style = ws_src_styles.cell(row=r, column=c)
        
        # Obtener valor
        if ws_src_values is not None:
            try:
                raw_value = ws_src_values.cell(row=r, column=c).value
            except Exception:
                raw_value = cell_style.value
        else:
            raw_value = cell_style.value

        # Detectar formato original
        orig_fmt = None
        try:
            orig_fmt = cell_style.number_format
        except Exception:
            orig_fmt = None
        custom_fmt = is_custom_number_format(orig_fmt)

        # Manejo especial para porcentajes en texto
        if isinstance(raw_value, str) and '%' in raw_value:
            try:
                cleaned = raw_value.replace('%', '').strip()
                parsed_num, _ = try_parse_number(cleaned)
                if parsed_num is not None:
                    final_val = parsed_num / 100.0 if isinstance(parsed_num, (int, float)) else None
                    if final_val is not None:
                        tgt = ws_tgt.cell(row=r, column=c, value=final_val)
                        copy_cell_style(cell_style, tgt)
                        if custom_fmt:
                            try:
                                if orig_fmt:
                                    tgt.number_format = orig_fmt
                            except Exception:
                                pass
                        else:
                            try:
                                tgt.number_format = '0.00%'
                            except Exception:
                                pass
                        return
            except Exception:
                pass

        # Si es string: limpiar y reintentar parseo
        if isinstance(raw_value, str):
            val = clean_text(raw_value)

            # Intentar parsear como fecha con mes en inglés primero
            parsed_date = try_parse_date(val)
            if parsed_date is not None:
                tgt = ws_tgt.cell(row=r, column=c, value=parsed_date)
                copy_cell_style(cell_style, tgt)
                if not custom_fmt:
                    try:
                        tgt.number_format = 'dd-mm-yyyy'  
                    except Exception:
                        pass
                else:
                    try:
                        if orig_fmt:
                            tgt.number_format = orig_fmt
                    except Exception:
                        pass
                return

            parsed_num, tag = try_parse_number(val)
            if parsed_num is not None:
                # Si el parser indicó 'AS_TEXT' -> escribir como texto para preservar cada dígito
                if tag == 'AS_TEXT':
                    # Modo seguro por defecto (OPCIÓN A): PRESERVAR como texto pero mejorar apariencia
                    text_to_write = str(parsed_num)
                    tgt = ws_tgt.cell(row=r, column=c, value=text_to_write)
                    copy_cell_style(cell_style, tgt)
                    try:
                        tgt.number_format = '@'  # forzar formato texto
                    except Exception:
                        pass
                    try:
                        tgt.alignment = Alignment(horizontal='right')  # parecer número (alineado a la derecha)
                    except Exception:
                        pass
                    return
                else:
                    # tag == 'General' -> escribir como número (int/float seguro)
                    tgt = ws_tgt.cell(row=r, column=c, value=parsed_num)
                    copy_cell_style(cell_style, tgt)
                    if not custom_fmt:
                        try:
                            if isinstance(parsed_num, float) and parsed_num.is_integer():
                                tgt.number_format = '0'
                            elif isinstance(parsed_num, int):
                                tgt.number_format = '0'
                            else:
                                tgt.number_format = '0.00'
                        except Exception:
                            pass
                    else:
                        try:
                            if orig_fmt:
                                tgt.number_format = orig_fmt
                        except Exception:
                            pass
                    return

            tgt = ws_tgt.cell(row=r, column=c, value=val)
            copy_cell_style(cell_style, tgt)
            if not custom_fmt:
                try:
                    tgt.number_format = '@'
                except Exception:
                    pass
            else:
                try:
                    if orig_fmt:
                        tgt.number_format = orig_fmt
                except Exception:
                    pass
            return

        # Si no es string: escribir tal cual
        tgt = ws_tgt.cell(row=r, column=c, value=raw_value)
        copy_cell_style(cell_style, tgt)

        if custom_fmt:
            try:
                if orig_fmt:
                    tgt.number_format = orig_fmt
            except Exception:
                pass
        else:
            try:
                if isinstance(raw_value, int):
                    tgt.number_format = '0'
                elif isinstance(raw_value, float):
                    try:
                        if raw_value.is_integer():
                            tgt.number_format = '0'
                        else:
                            tgt.number_format = '0.00'
                    except Exception:
                        tgt.number_format = '0.00'
                elif isinstance(raw_value, datetime):
                    tgt.number_format = 'dd-mm-yyyy' 
                elif raw_value is None:
                    pass
                else:
                    tgt.number_format = '@'
            except Exception:
                pass
                
    except Exception as e:
        logger.debug(f"Error en process_cell({r},{c}): {e}")
        # Escribir al menos valor vacío para no romper estructura
        try:
            ws_tgt.cell(row=r, column=c, value=None)
        except:
            pass

# ---------------------------
# Generar nombre único
# ---------------------------
def _unique_out_path(dest_folder: Path, name: str):
    """Genera ruta única para archivo de salida"""
    try:
        base = Path(name).stem
        ext = Path(name).suffix or ".xlsx"
        out = dest_folder / (base + ext)
        if not out.exists():
            return out
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        newname = f"{base}_{ts}{ext}"
        return dest_folder / newname
    except Exception as e:
        logger.error(f"Error generando ruta única: {e}")
        raise

# ---------------------------
# Verificar tamaño de archivo
# ---------------------------
def check_file_size(filepath):
    """Verifica tamaño de archivo y retorna advertencia si es necesario"""
    try:
        size_mb = os.path.getsize(filepath) / (1024 * 1024)
        logger.info(f"Tamaño de archivo: {size_mb:.2f} MB")
        return size_mb
    except Exception as e:
        logger.warning(f"No se pudo verificar tamaño de archivo: {e}")
        return 0

# ---------------------------
# Variante xlwings in-place
# ---------------------------
def process_workbook_xlwings_inplace(filepath, dest_folder):
    """Procesa con xlwings de manera robusta"""
    logger.info(f"Iniciando proceso xlwings para: {filepath}")
    
    if not USE_XLWINGS:
        raise RuntimeError("xlwings no está disponible")

    path = Path(filepath)
    dest_folder = Path(dest_folder)
    dest_folder.mkdir(parents=True, exist_ok=True)
    out_path = _unique_out_path(dest_folder, path.name)

    app = None
    wb = None
    
    try:
        # Verificar tamaño
        size_mb = check_file_size(filepath)
        if size_mb > MAX_FILE_SIZE_MB:
            logger.warning(f"Archivo grande detectado: {size_mb:.2f} MB")
        
        # Copiar archivo
        logger.info(f"Copiando archivo a: {out_path}")
        shutil.copy2(str(path), str(out_path))

        # Abrir con xlwings
        logger.info("Abriendo con xlwings...")
        app = xw.App(visible=False, add_book=False)
        wb = app.books.open(str(out_path))

        # Procesar hojas
        total_sheets = len(wb.sheets)
        logger.info(f"Procesando {total_sheets} hojas...")
        
        for idx, sh in enumerate(wb.sheets, 1):
            try:
                logger.info(f"Procesando hoja {idx}/{total_sheets}: {sh.name}")
                rng = sh.used_range
                vals = rng.value
                rng.value = vals
            except Exception as e:
                logger.warning(f"Error en hoja {sh.name}: {e}")
                try:
                    used = sh.api.UsedRange
                    rng2 = sh.range(used.Address)
                    vals2 = rng2.value
                    rng2.value = vals2
                except Exception as e2:
                    logger.error(f"Error crítico en hoja {sh.name}: {e2}")

        logger.info("Guardando archivo...")
        wb.save()
        wb.close()
        app.quit()
        
        logger.info(f"Proceso xlwings completado: {out_path}")
        return str(out_path)
        
    except Exception as e:
        logger.error(f"Error en proceso xlwings: {e}")
        logger.error(traceback.format_exc())
        
        # Limpieza
        try:
            if wb:
                wb.close()
        except:
            pass
        try:
            if app:
                app.quit()
        except:
            pass
        
        raise RuntimeError(f"Error al procesar con xlwings: {e}")

# ---------------------------
# Proceso principal openpyxl
# ---------------------------
def process_workbook_openpyxl_copy(filepath, dest_folder):
    """Procesa con openpyxl de manera robusta"""
    logger.info(f"Iniciando proceso openpyxl para: {filepath}")
    
    path = Path(filepath)
    dest_folder = Path(dest_folder)
    dest_folder.mkdir(parents=True, exist_ok=True)
    out_path = _unique_out_path(dest_folder, path.name)

    try:
        # Verificar tamaño
        size_mb = check_file_size(filepath)
        if size_mb > MAX_FILE_SIZE_MB:
            logger.warning(f"Archivo grande detectado: {size_mb:.2f} MB. El proceso puede tardar.")

        # Intentar obtener valores con xlwings si está disponible (opcional)
        values_by_sheet = {}
        used_xlwings_for_values = False
        
        if USE_XLWINGS and size_mb < MAX_FILE_SIZE_MB:
            try:
                logger.info("Intentando obtener valores con xlwings...")
                app = xw.App(visible=False, add_book=False)
                wb_x = app.books.open(str(path))
                for sh in wb_x.sheets:
                    try:
                        vals = sh.used_range.value
                        values_by_sheet[sh.name] = vals
                    except Exception as e:
                        logger.warning(f"Error obteniendo valores de hoja {sh.name}: {e}")
                wb_x.close()
                app.quit()
                used_xlwings_for_values = True
                logger.info("Valores obtenidos con xlwings exitosamente")
            except Exception as e:
                logger.warning(f"No se pudo usar xlwings para valores: {e}")
                try:
                    app.quit()
                except:
                    pass
                used_xlwings_for_values = False

        # Cargar workbooks
        logger.info("Cargando workbook con estilos...")
        try:
            wb_styles = load_workbook(filename=str(path), data_only=False, read_only=False)
        except Exception as e:
            logger.error(f"Error cargando workbook: {e}")
            raise RuntimeError(f"No se pudo abrir {path} con openpyxl: {e}")

        wb_values = None
        if used_xlwings_for_values:
            logger.info("Creando workbook de valores desde xlwings...")
            wb_values = Workbook()
            if wb_values.worksheets:
                wb_values.remove(wb_values.worksheets[0])
            
            for sheetname in wb_styles.sheetnames:
                vals = values_by_sheet.get(sheetname, None)
                ws_v = wb_values.create_sheet(title=sheetname)
                if vals is None:
                    continue
                if not isinstance(vals, list):
                    vals = [[vals]]
                if all(not isinstance(r, list) for r in vals):
                    vals = [vals]
                    
                for r_idx, row in enumerate(vals, start=1):
                    for c_idx, cell_val in enumerate(row, start=1):
                        try:
                            if isinstance(cell_val, str):
                                cell_val = clean_text(cell_val)
                                # Intentar parsear como fecha primero
                                parsed_date = try_parse_date(cell_val)
                                if parsed_date is not None:
                                    ws_v.cell(row=r_idx, column=c_idx, value=parsed_date)
                                    ws_v.cell(row=r_idx, column=c_idx).number_format = 'dd-mm-yyyy'  
                                    continue
                                    
                                parsed_num, tag = try_parse_number(cell_val)
                                if parsed_num is not None and tag == 'General':
                                    ws_v.cell(row=r_idx, column=c_idx, value=parsed_num)
                                    continue
                                if parsed_num is not None and tag == 'AS_TEXT':
                                    ws_v.cell(row=r_idx, column=c_idx, value=str(parsed_num))
                                    continue
                            ws_v.cell(row=r_idx, column=c_idx, value=cell_val)
                        except Exception as e:
                            logger.debug(f"Error procesando valor ({r_idx},{c_idx}): {e}")
        else:
            logger.info("Cargando workbook con valores (data_only)...")
            try:
                wb_values = load_workbook(filename=str(path), data_only=True, read_only=False)
            except Exception as e:
                logger.warning(f"No se pudo cargar workbook con data_only: {e}")
                wb_values = None

        # Crear nuevo workbook
        logger.info("Creando nuevo workbook...")
        new_wb = Workbook()
        if new_wb.worksheets:
            new_wb.remove(new_wb.worksheets[0])

        # Procesar hojas
        total_sheets = len(wb_styles.sheetnames)
        for idx, sheetname in enumerate(wb_styles.sheetnames, 1):
            try:
                logger.info(f"Procesando hoja {idx}/{total_sheets}: {sheetname}")
                ws_src_styles = wb_styles[sheetname]
                ws_src_values = None
                if wb_values is not None and sheetname in wb_values.sheetnames:
                    ws_src_values = wb_values[sheetname]
                ws_tgt = new_wb.create_sheet(title=sheetname)
                copy_sheet_visuals(ws_src_styles, ws_src_values, ws_tgt)
                
                # Liberar memoria
                gc.collect()
                
            except Exception as e:
                logger.error(f"Error procesando hoja {sheetname}: {e}")
                logger.error(traceback.format_exc())
                # Crear hoja vacía para no romper el archivo
                try:
                    if sheetname not in new_wb.sheetnames:
                        new_wb.create_sheet(title=sheetname)
                except:
                    pass

        # Guardar
        logger.info(f"Guardando archivo en: {out_path}")
        new_wb.save(out_path)
        
        # Cerrar workbooks explícitamente
        try:
            wb_styles.close()
            if wb_values:
                wb_values.close()
            new_wb.close()
        except:
            pass
        
        # Liberar memoria
        gc.collect()
        
        logger.info(f"Proceso openpyxl completado: {out_path}")
        return str(out_path)
        
    except Exception as e:
        logger.error(f"Error crítico en proceso openpyxl: {e}")
        logger.error(traceback.format_exc())
        raise

# ---------------------------
# Wrapper principal
# ---------------------------
def process_workbook(filepath, dest_folder, use_xlwings_mode=False):
    """Wrapper principal con manejo robusto de errores"""
    try:
        logger.info(f"{'='*60}")
        logger.info(f"Iniciando procesamiento: {filepath}")
        logger.info(f"Modo xlwings: {use_xlwings_mode}")
        
        if use_xlwings_mode:
            if not USE_XLWINGS:
                raise RuntimeError("xlwings no está disponible")
            return process_workbook_xlwings_inplace(filepath, dest_folder)
        else:
            return process_workbook_openpyxl_copy(filepath, dest_folder)
            
    except Exception as e:
        logger.error(f"Error procesando {filepath}: {e}")
        logger.error(traceback.format_exc())
        raise

# ---------------------------
# INTERFAZ GUI (sin cambios funcionales relevantes)
# ---------------------------
def preguntar_carpeta_destino_var(carpeta_root):
    """Pregunta dónde guardar archivos con manejo de errores"""
    try:
        opciones_msg = (
            "¿Dónde deseas guardar los archivos mejorados?\n\n"
            "Opciones:\n"
            "1. Guardar en carpeta raíz (limpios)\n"
            "2. Seleccionar subcarpeta existente\n"
            "3. Crear nueva subcarpeta\n\n"
            "Escribe 1, 2 o 3:"
        )
        seleccion = simpledialog.askstring("Destino de archivos", opciones_msg)
        
        if seleccion == "2":
            subcarpetas = [d for d in os.listdir(carpeta_root) if os.path.isdir(os.path.join(carpeta_root, d))]
            if not subcarpetas:
                messagebox.showinfo("Sin subcarpetas", "No hay subcarpetas disponibles. Se usará la carpeta raíz.")
                return carpeta_root
            subcarpeta = simpledialog.askstring("Elegir subcarpeta", f"Subcarpetas disponibles:\n\n" + "\n".join(subcarpetas) + "\n\nEscribe el nombre exacto:")
            if subcarpeta and subcarpeta in subcarpetas:
                return os.path.join(carpeta_root, subcarpeta)
            else:
                messagebox.showinfo("Error", "Nombre inválido. Se usará la carpeta raíz.")
                return carpeta_root
        elif seleccion == "3":
            nueva = simpledialog.askstring("Crear subcarpeta", "Escribe el nombre de la nueva subcarpeta:")
            if nueva:
                ruta_nueva = os.path.join(carpeta_root, nueva)
                os.makedirs(ruta_nueva, exist_ok=True)
                logger.info(f"Subcarpeta creada: {ruta_nueva}")
                return ruta_nueva
            else:
                return carpeta_root
        else:
            return carpeta_root
    except Exception as e:
        logger.error(f"Error en preguntar_carpeta_destino_var: {e}")
        return carpeta_root

def convertir_xls_a_xlsx_si_necesario(path):
    """Placeholder para conversión XLS (no implementado)"""
    return path

def procesar_carpeta(carpeta, use_xlwings_mode=False):
    """Procesa carpeta completa con manejo robusto de errores"""
    try:
        logger.info(f"Iniciando procesamiento de carpeta: {carpeta}")
        carpeta = str(carpeta)
        carpeta_destino = preguntar_carpeta_destino_var(CARPETA_LIMPIOS)
        
        archivos_validos = []
        try:
            archivos_validos = [f for f in os.listdir(carpeta) if f.lower().endswith((".xlsx", ".xlsm", ".xls"))]
            logger.info(f"Archivos encontrados: {len(archivos_validos)}")
        except Exception as e:
            logger.error(f"Error listando archivos: {e}")
            messagebox.showerror("Error", f"No se pudo leer la carpeta: {e}")
            return

        if not archivos_validos:
            messagebox.showinfo("Sin archivos", "No se encontraron archivos Excel en la carpeta seleccionada.")
            logger.warning("No se encontraron archivos Excel")
            return

        if len(archivos_validos) > 1000:
            archivos_validos = archivos_validos[:1000]
            messagebox.showinfo("Límite aplicado", "Solo se procesarán los primeros 1000 archivos.")
            logger.warning("Límite de 1000 archivos aplicado")

        # Verificar posibles sobrescrituras
        nombres_existentes = []
        for archivo in archivos_validos:
            nombre = Path(archivo).stem + Path(archivo).suffix
            ruta_out = Path(carpeta_destino) / nombre
            if ruta_out.exists():
                nombres_existentes.append(nombre)

        if nombres_existentes:
            lista = "\n".join(nombres_existentes[:5]) + ("\n..." if len(nombres_existentes) > 5 else "")
            respuesta = messagebox.askyesno(
                "Advertencia de sobrescritura", 
                f"⚠️ Estos archivos ya existen en destino (se guardarán con sufijo fecha/hora):\n\n{lista}\n\n¿Deseas continuar?"
            )
            if not respuesta:
                logger.info("Usuario canceló el proceso")
                return

        total = len(archivos_validos)
        exitosos = 0
        errores = 0
        archivos_con_error = []
        
        progreso["maximum"] = total
        progreso["value"] = 0
        root.update_idletasks()

        logger.info(f"Iniciando procesamiento de {total} archivos...")

        for i, archivo in enumerate(archivos_validos, start=1):
            ruta = os.path.join(carpeta, archivo)
            try:
                logger.info(f"[{i}/{total}] Procesando: {archivo}")
                etiqueta_estado_var.set(f"Procesando [{i}/{total}]: {archivo}")
                root.update_idletasks()
                
                # Verificar tamaño antes de procesar
                size_mb = check_file_size(ruta)
                if size_mb > MAX_FILE_SIZE_MB * 2:  # Más del doble del límite recomendado
                    logger.warning(f"Archivo muy grande: {size_mb:.2f} MB")
                    respuesta = messagebox.askyesno(
                        "Archivo grande detectado",
                        f"El archivo '{archivo}' es muy grande ({size_mb:.2f} MB).\n"
                        f"El procesamiento puede tardar varios minutos.\n\n"
                        f"¿Deseas continuar con este archivo?"
                    )
                    if not respuesta:
                        logger.info(f"Usuario omitió archivo grande: {archivo}")
                        etiqueta_estado_var.set(f"Omitido: {archivo}")
                        continue
                
                ruta_proc = convertir_xls_a_xlsx_si_necesario(ruta)
                out = process_workbook(ruta_proc, carpeta_destino, use_xlwings_mode=use_xlwings_mode)
                
                exitosos += 1
                progreso["value"] = i
                etiqueta_estado_var.set(f"✓ Completado [{i}/{total}]: {Path(out).name}")
                logger.info(f"✓ Archivo procesado exitosamente: {out}")
                
            except Exception as e:
                errores += 1
                archivos_con_error.append(f"{archivo}: {str(e)[:100]}")
                logger.error(f"✗ Error procesando {archivo}: {e}")
                logger.error(traceback.format_exc())
                etiqueta_estado_var.set(f"✗ Error [{i}/{total}]: {archivo}")
                
            finally:
                root.update_idletasks()
                # Liberar memoria cada 10 archivos
                if i % 10 == 0:
                    gc.collect()

        # Resumen final
        mensaje_final = f"Proceso completado:\n\n"
        mensaje_final += f"✓ Exitosos: {exitosos}\n"
        mensaje_final += f"✗ Errores: {errores}\n"
        mensaje_final += f"Total: {total}\n\n"
        mensaje_final += f"Archivos guardados en:\n{carpeta_destino}\n\n"
        
        if archivos_con_error:
            mensaje_final += f"Archivos con errores:\n"
            for error_info in archivos_con_error[:5]:
                mensaje_final += f"• {error_info}\n"
            if len(archivos_con_error) > 5:
                mensaje_final += f"... y {len(archivos_con_error) - 5} más.\n"
            mensaje_final += f"\nRevisa el log para más detalles:\n{log_filename}"
        
        logger.info(f"Proceso finalizado. Exitosos: {exitosos}, Errores: {errores}")
        
        if errores > 0:
            messagebox.showwarning("Proceso completado con errores", mensaje_final)
        else:
            messagebox.showinfo("Proceso completado", mensaje_final)
            
    except Exception as e:
        logger.error(f"Error crítico en procesar_carpeta: {e}")
        logger.error(traceback.format_exc())
        messagebox.showerror("Error crítico", f"Error inesperado:\n{e}\n\nRevisa el log: {log_filename}")

def seleccionar_carpeta(use_xlwings_mode=False):
    """Selecciona carpeta con validación"""
    try:
        carpeta = filedialog.askdirectory(title="Selecciona carpeta con archivos Excel")
        if carpeta:
            logger.info(f"Carpeta seleccionada: {carpeta}")
            procesar_carpeta(carpeta, use_xlwings_mode=use_xlwings_mode)
        else:
            logger.info("Usuario canceló selección de carpeta")
    except Exception as e:
        logger.error(f"Error en seleccionar_carpeta: {e}")
        messagebox.showerror("Error", f"Error al seleccionar carpeta: {e}")

def abrir_carpeta_limpios():
    """Abre carpeta de archivos limpios"""
    try:
        path_absoluto = os.path.join(
            os.path.dirname(sys.executable if getattr(sys, 'frozen', False) else __file__), 
            CARPETA_LIMPIOS
        )
        if not os.path.exists(path_absoluto):
            os.makedirs(path_absoluto, exist_ok=True)
            logger.info(f"Carpeta creada: {path_absoluto}")
        
        logger.info(f"Abriendo carpeta: {path_absoluto}")
        
        if sys.platform.startswith('win'):
            os.startfile(path_absoluto)
        elif sys.platform == 'darwin':
            subprocess.run(["open", path_absoluto])
        else:
            subprocess.run(["xdg-open", path_absoluto])
            
    except Exception as e:
        logger.error(f"Error abriendo carpeta limpios: {e}")
        messagebox.showerror("Error", f"No se pudo abrir la carpeta: {e}\n\nRuta: {path_absoluto}")

def abrir_carpeta_logs():
    """Abre carpeta de logs"""
    try:
        path_absoluto = os.path.abspath(LOG_DIR)
        if not os.path.exists(path_absoluto):
            os.makedirs(path_absoluto, exist_ok=True)
        
        logger.info(f"Abriendo carpeta de logs: {path_absoluto}")
        
        if sys.platform.startswith('win'):
            os.startfile(path_absoluto)
        elif sys.platform == 'darwin':
            subprocess.run(["open", path_absoluto])
        else:
            subprocess.run(["xdg-open", path_absoluto])
            
    except Exception as e:
        logger.error(f"Error abriendo carpeta logs: {e}")
        messagebox.showerror("Error", f"No se pudo abrir la carpeta de logs: {e}")

def mostrar_instrucciones():
    """Muestra instrucciones de uso"""
    try:
        instrucciones = (
            "📄 Instrucciones de uso:\n\n"
            "1. Haz clic en 'Seleccionar Carpeta con Archivos'.\n"
            "2. Elige la carpeta que contiene los archivos Excel (.xlsx, .xlsm, .xls).\n"
            "3. Se te pedirá dónde guardar los archivos mejorados (raíz / subcarpeta / crear nueva).\n"
            "4. El programa procesará los archivos y guardará copias con fórmulas reemplazadas por valores.\n"
            "   - Si un archivo destino ya existe, se guardará con sufijo _YYYYMMDD_HHMMSS.\n"
            "5. Modo xlwings (in-place): si activas la casilla y xlwings/Excel están disponibles,\n"
            "   se creará una copia y Excel reemplazará todas las fórmulas por valores.\n"
            "   Esto preserva macros, gráficos, imágenes y cualquier objeto exactamente como en el original.\n"
            "6. Se respetarán exactamente las filas/columnas inmovilizadas (freeze panes) del archivo original.\n"
            "7. Si xlwings no está disponible se usará el método openpyxl por defecto.\n"
            "8. No interrumpas el proceso hasta que finalice.\n"
            "9. Revisa los logs en caso de errores.\n\n"
            f"📊 Log actual: {log_filename}\n"
            f"📁 Carpeta de logs: {LOG_DIR}\n"
        )
        messagebox.showinfo("Instrucciones de Uso", instrucciones)
    except Exception as e:
        logger.error(f"Error mostrando instrucciones: {e}")

def mostrar_info_sistema():
    """Muestra información del sistema y configuración"""
    try:
        info = "ℹ️ Información del Sistema\n\n"
        info += f"Python: {sys.version.split()[0]}\n"
        info += f"openpyxl: Instalado ✓\n"
        info += f"python-dateutil: Instalado ✓\n"
        info += f"xlwings: {'Instalado ✓' if USE_XLWINGS else 'No disponible ✗'}\n\n"
        info += f"📁 Carpeta base: {BASE_DIR}\n"
        info += f"📁 Carpeta limpios: {CARPETA_LIMPIOS}\n"
        info += f"📁 Carpeta logs: {LOG_DIR}\n"
        info += f"📄 Log actual: {os.path.basename(log_filename)}\n\n"
        info += f"⚙️ Configuración:\n"
        info += f"• Tamaño máx recomendado: {MAX_FILE_SIZE_MB} MB\n"
        info += f"• Procesamiento por chunks: {CHUNK_SIZE} filas\n        \n        "
        
        messagebox.showinfo("Información del Sistema", info)
    except Exception as e:
        logger.error(f"Error mostrando info sistema: {e}")

# ---------------------------
# Construir UI
# ---------------------------
try:
    root = tk.Tk()
    root.title("Limpiador de Excel por Lotes    By: Erick")
    root.geometry("600x480")
    root.resizable(False, False)

    frame = ttk.Frame(root, padding=18)
    frame.pack(expand=True, fill='both')

    # Título
    ttk.Label(frame, text="Herramienta: Limpiar & Convertir Excel", font=("Segoe UI", 14, "bold")).pack(pady=8)
    
    # Separador
    ttk.Separator(frame, orient='horizontal').pack(fill='x', pady=5)

    # Checkbox xlwings
    use_xlwings_var = tk.BooleanVar(value=False)
    chk_text = "✨ Procesar con Excel (xlwings) — copia + reemplazar fórmulas"
    chk = ttk.Checkbutton(frame, text=chk_text, variable=use_xlwings_var)
    chk.pack(pady=6, fill='x')

    if not USE_XLWINGS:
        chk.state(['disabled'])
        lbl_xlw_hint = ttk.Label(
            frame, 
            text="⚠️ xlwings no detectado. Instala: pip install xlwings",
            foreground="orange"
        )
        lbl_xlw_hint.pack(pady=(0,6))

    # Separador
    ttk.Separator(frame, orient='horizontal').pack(fill='x', pady=8)

    # Botón principal
    btn_select = ttk.Button(
        frame,
        text="📂 Seleccionar Carpeta con Archivos",
        command=lambda: seleccionar_carpeta(use_xlwings_mode=use_xlwings_var.get())
    )
    btn_select.pack(pady=8, fill='x', ipady=5)

    # Barra de progreso
    progreso = ttk.Progressbar(frame, length=400, mode='determinate')
    progreso.pack(pady=10, fill='x')

    # Estado
    etiqueta_estado_var = tk.StringVar(value="⏳ Esperando acción...")
    lbl_estado = ttk.Label(frame, textvariable=etiqueta_estado_var, font=("Segoe UI", 9))
    lbl_estado.pack(pady=4)

    # Separador
    ttk.Separator(frame, orient='horizontal').pack(fill='x', pady=10)

    # Frame para botones adicionales
    frame_botones = ttk.Frame(frame)
    frame_botones.pack(fill='x', pady=5)

    ttk.Button(
        frame_botones, 
        text="📁 Ver archivos limpios", 
        command=abrir_carpeta_limpios
    ).pack(side='left', expand=True, fill='x', padx=2)

    ttk.Button(
        frame_botones, 
        text="📋 Ver logs", 
        command=abrir_carpeta_logs
    ).pack(side='left', expand=True, fill='x', padx=2)

    # Frame para botones de ayuda
    frame_ayuda = ttk.Frame(frame)
    frame_ayuda.pack(fill='x', pady=5)

    ttk.Button(
        frame_ayuda, 
        text="📘 Instrucciones", 
        command=mostrar_instrucciones
    ).pack(side='left', expand=True, fill='x', padx=2)

    ttk.Button(
        frame_ayuda, 
        text="ℹ️ Info Sistema", 
        command=mostrar_info_sistema
    ).pack(side='left', expand=True, fill='x', padx=2)

    # Separador
    ttk.Separator(frame, orient='horizontal').pack(fill='x', pady=8)

    # Nota final
    nota = ttk.Label(
        frame, 
        text="⚡ Todos los errores se registran en logs para diagnóstico",
        font=("Segoe UI", 8),
        foreground="gray"
    )
    nota.pack(pady=5)

    logger.info("Interfaz gráfica iniciada correctamente")
    logger.info(f"Log guardándose en: {log_filename}")
    
    root.mainloop()
    
except Exception as e:
    logger.critical(f"Error crítico al iniciar la aplicación: {e}")
    logger.critical(traceback.format_exc())
    try:
        messagebox.showerror(
            "Error Crítico", 
            f"No se pudo iniciar la aplicación:\n\n{e}\n\nRevisa el log: {log_filename}"
        )
    except:
        print(f"ERROR CRÍTICO: {e}")
    sys.exit(1)    