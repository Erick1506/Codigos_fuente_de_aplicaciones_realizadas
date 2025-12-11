#!/usr/bin/env python3
"""
Procesador automático de solicitudes de devolución (SENA) - Script completo con UI (Streamlit) y CLI.
Versión mejorada con detección robusta de fechas y documentos.
"""

import os
import sys
import re
import argparse
import hashlib
import subprocess
import logging
from datetime import datetime, timedelta
from dateutil.parser import parse as parse_date
import pytesseract
from pdf2image import convert_from_path
from PyPDF2 import PdfReader
import pandas as pd
import numpy as np
import cv2
from PIL import Image
import tempfile
import io
import unicodedata

# Excel templating
from openpyxl import load_workbook

# Streamlit import is optional for CLI mode
try:
    import streamlit as st
    STREAMLIT_AVAILABLE = True
except Exception:
    STREAMLIT_AVAILABLE = False

# optional fuzzy lib — use if installed
try:
    from rapidfuzz import fuzz
    HAVE_RAPIDFUZZ = True
except Exception:
    HAVE_RAPIDFUZZ = False

# ---------------- logging ----------------
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

# ---------------- CONFIG  ----------------
POPPLER_PATH = r"C:\poppler\Library\bin"
TESSERACT_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
TESSDATA_DIR = r"C:\Program Files\Tesseract-OCR\tessdata"  # carpeta que contiene los .traineddata

# Plantillas (colocar en la misma carpeta que el script o ajustar rutas)
TEMPLATE_MISIONAL = "Lista.xlsx"   # plantilla base misional (usuario puede reemplazar)
TEMPLATE_NO_MISIONAL = "lista2.xlsx"  # plantilla base no_misional (usuario puede reemplazar)

# Aplicar configuración de pytesseract y TESSDATA_PREFIX
os.environ.setdefault("TESSDATA_PREFIX", TESSDATA_DIR)
pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

# Verificar que tesseract y spa estén disponibles (advertencia si faltas)
def verificar_tesseract_y_idioma(idioma_req="spa"):
    try:
        proc = subprocess.run([TESSERACT_CMD, "--list-langs"], capture_output=True, text=True, check=True)
        out_lines = [ln.strip() for ln in proc.stdout.splitlines() if ln.strip()]
        if len(out_lines) > 1 and out_lines[0].lower().startswith("list of available languages"):
            installed = [ln.strip().lower() for ln in out_lines[1:]]
        else:
            installed = [ln.strip().lower() for ln in out_lines]
        logging.info(f"Tesseract ejecutable: {TESSERACT_CMD}")
        logging.info(f"tessdata (TESSDATA_PREFIX) apuntando a: {os.environ.get('TESSDATA_PREFIX')}")
        logging.info(f"Idiomas instalados en tessdata: {installed}")
        if idioma_req.lower() not in installed:
            logging.warning(f"Idioma requerido '{idioma_req}' no encontrado en tessdata (detectados: {installed}). "
                            f"Considera instalar '{idioma_req}.traineddata' en {TESSDATA_DIR}.")
    except subprocess.CalledProcessError as e:
        logging.warning(f"Error al invocar Tesseract para listar idiomas: {e.stderr or e.stdout}")
    except Exception as e:
        logging.warning(f"No se pudo verificar Tesseract/tessdata: {e}")

verificar_tesseract_y_idioma("spa")

# ---------------- festivos y utilidades ----------------
COLOMBIA_FESTIVOS = {
    "2025-01-01", "2025-05-01", "2025-07-20", "2025-08-07", "2025-12-08", "2025-12-25"
}

def sha256_file(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()

def es_dia_habil(fecha: datetime.date):
    if fecha.weekday() >= 5:
        return False
    return fecha.strftime("%Y-%m-%d") not in COLOMBIA_FESTIVOS

def fecha_limite_habiles(fecha_expedicion: datetime, dias_habiles: int):
    fecha = fecha_expedicion
    contador = 0
    while contador < dias_habiles:
        fecha = fecha + timedelta(days=1)
        if es_dia_habil(fecha.date()):
            contador += 1
    return fecha

# ---------------- OCR / imagen ----------------
def ocr_image_and_confidence(pil_image: Image.Image, lang='spa'):
    try:
        data = pytesseract.image_to_data(pil_image, lang=lang, output_type=pytesseract.Output.DICT)
    except Exception:
        text = pytesseract.image_to_string(pil_image, lang=lang)
        return text, 0.0
    words = data.get('text', [])
    confs = data.get('conf', [])
    text = " ".join([w for w in words if w and w.strip()])
    conf_vals = []
    for c in confs:
        try:
            ci = float(c)
            if ci >= 0:
                conf_vals.append(ci)
        except:
            pass
    avg_conf = float(np.mean(conf_vals)) if conf_vals else 0.0
    return text, avg_conf

def detectar_firma_manuscrita(pil_image: Image.Image, thresh=0.03):
    img = np.array(pil_image.convert("L"))
    h, w = img.shape
    roi = img[int(h*0.62):h, 0:w]
    _, bw = cv2.threshold(roi, 200, 255, cv2.THRESH_BINARY_INV)
    ratio = (bw > 0).sum() / (bw.size + 1e-9)
    return ratio > thresh

# ---------------- MEJORADO: Expresiones regulares para fechas ----------------
# Patrones más amplios para capturar diversos formatos de fecha

# Patrón principal para fechas numéricas (muy permisivo)
RE_DATE_NUMERIC = re.compile(
    r'\b(\d{1,2})[/\-\.\s\\]+(\d{1,2})[/\-\.\s\\]+(\d{2,4})\b|'  # DD/MM/YYYY o DD-MM-YYYY
    r'\b(\d{2,4})[/\-\.\s\\]+(\d{1,2})[/\-\.\s\\]+(\d{1,2})\b',  # YYYY/MM/DD o YYYY-MM-DD
    re.IGNORECASE
)

# Patrón para fechas con nombre de mes
RE_DATE_TEXTO = re.compile(
    r'\b(\d{1,2})\s*(?:de\s+)?([A-Za-zÁÉÍÓÚáéíóúñÑ]+)\s*(?:de\s+|del?\s+)?(\d{2,4})\b',
    re.IGNORECASE
)

# Patrones específicos para contexto bancario (busca "expedición", "fecha", etc.)
RE_DATE_BANCARIA_CONTEXTO = re.compile(
    r'(?:fecha|expedici[oó]n|expide|emitid[ao]|generad[ao]|cread[ao]|realizad[ao]|elabor[aó]|elaboraci[oó]n)'
    r'[\s\:\-\.\,]*'
    r'(\d{1,2}[/\-\.\s\\]+\d{1,2}[/\-\.\s\\]+\d{2,4}|'
    r'\d{1,2}\s+(?:de\s+)?[A-Za-zÁÉÍÓÚáéíóúñÑ]+\s+(?:de\s+|del?\s+)?\d{2,4})',
    re.IGNORECASE
)

# Patrón de compatibilidad (original)
RE_DATE = re.compile(
    r'\b(\d{1,2})[/\-\.\s](\d{1,2})[/\-\.\s](\d{2,4})\b|'
    r'\b(\d{2,4})[/\-\.\s](\d{1,2})[/\-\.\s](\d{1,2})\b|'
    r'\b(\d{1,2})\s*(de|/)\s*([A-Za-z]+)\s*(de|/)\s*(\d{2,4})\b',
    re.IGNORECASE
)

# Expresión regular específica para fechas en documentos bancarios (compatibilidad)
RE_DATE_BANCARIA = re.compile(
    r'(fecha|expedici[oó]n|expide|emitid[ao]|generad[ao]|cread[ao]|realizad[ao])[\s\:\-]*'
    r'(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4}|\d{1,2}\s+de\s+[A-Za-z]+\s+de\s+\d{4})',
    re.IGNORECASE
)

# Mapeo expandido de meses en español (incluyendo variaciones)
MESES_ESPANOL = {
    'enero': 1, 'ene': 1, 'ener': 1,
    'febrero': 2, 'feb': 2, 'febr': 2,
    'marzo': 3, 'mar': 3, 'mzo': 3,
    'abril': 4, 'abr': 4, 'abrl': 4,
    'mayo': 5, 'may': 5,
    'junio': 6, 'jun': 6,
    'julio': 7, 'jul': 7,
    'agosto': 8, 'ago': 8, 'agos': 8, 'agto': 8,
    'septiembre': 9, 'sep': 9, 'sept': 9, 'setiembre': 9, 'set': 9,
    'octubre': 10, 'oct': 10,
    'noviembre': 11, 'nov': 11,
    'diciembre': 12, 'dic': 12, 'dbre': 12
}

# ---------------- MEJORADO: Funciones para extracción de fechas ----------------
def limpiar_texto_fecha(texto: str) -> str:
    """
    Limpia el texto para mejorar la detección de fechas.
    Maneja errores comunes de OCR.
    """
    if not texto:
        return ""
    
    # Reemplazos comunes de errores OCR
    texto = texto.replace('|', '/')
    texto = re.sub(r'\bl\b', '1', texto)  # l minúscula aislada por 1
    texto = re.sub(r'(?<=\d)O(?=\d)', '0', texto)  # O entre números
    texto = re.sub(r'(?<=\d)o(?=\d)', '0', texto)  # o entre números
    texto = re.sub(r'[\[\]\{\}\(\)]', '', texto)  # quitar brackets
    
    return texto

def extraer_todas_fechas_texto(texto: str, contexto_bancario=False):
    """
    Extrae TODAS las posibles fechas del texto usando múltiples estrategias.
    Devuelve una lista de fechas encontradas ordenadas.
    """
    fechas_encontradas = []
    
    # Limpiar texto
    texto_limpio = limpiar_texto_fecha(texto)
    
    # ESTRATEGIA 1: Buscar fechas con contexto bancario específico
    if contexto_bancario:
        matches_contexto = RE_DATE_BANCARIA_CONTEXTO.finditer(texto_limpio)
        for match in matches_contexto:
            fecha_texto = match.group(1)
            fecha = parsear_fecha_flexible(fecha_texto)
            if fecha:
                fechas_encontradas.append(fecha)
                logging.debug(f"Fecha encontrada con contexto bancario: {fecha} de '{fecha_texto}'")
    
    # ESTRATEGIA 2: Buscar fechas con nombre de mes
    matches_texto = RE_DATE_TEXTO.finditer(texto_limpio)
    for match in matches_texto:
        try:
            dia = int(match.group(1))
            mes_nombre = match.group(2).lower().strip()
            año = int(match.group(3))
            
            # Ajustar año de 2 dígitos
            if año < 100:
                año += 2000 if año < 50 else 1900
            
            # Buscar mes en el diccionario (fuzzy match)
            mes = None
            for key, value in MESES_ESPANOL.items():
                if mes_nombre.startswith(key[:3]) or key.startswith(mes_nombre[:3]):
                    mes = value
                    break
            
            if mes and 1 <= dia <= 31 and 1900 <= año <= 2100:
                fecha = datetime(año, mes, dia).date()
                fechas_encontradas.append(fecha)
                logging.debug(f"Fecha encontrada con texto: {fecha} de '{match.group(0)}'")
        except (ValueError, AttributeError) as e:
            logging.debug(f"Error parseando fecha con texto: {e}")
            continue
    
    # ESTRATEGIA 3: Buscar fechas numéricas puras
    matches_numeric = RE_DATE_NUMERIC.finditer(texto_limpio)
    for match in matches_numeric:
        # Probar diferentes grupos de captura
        for i in range(1, 7, 3):
            if match.group(i) and match.group(i+1) and match.group(i+2):
                try:
                    num1 = int(match.group(i))
                    num2 = int(match.group(i+1))
                    num3 = int(match.group(i+2))
                    
                    # Ajustar año si es de 2 dígitos
                    if num3 < 100:
                        num3 += 2000 if num3 < 50 else 1900
                    
                    # Intentar DD/MM/YYYY primero (formato latinoamericano)
                    if 1 <= num1 <= 31 and 1 <= num2 <= 12 and 1900 <= num3 <= 2100:
                        fecha = datetime(num3, num2, num1).date()
                        fechas_encontradas.append(fecha)
                        logging.debug(f"Fecha encontrada numérica (DD/MM/YYYY): {fecha} de '{match.group(0)}'")
                    # Intentar YYYY/MM/DD
                    elif 1900 <= num1 <= 2100 and 1 <= num2 <= 12 and 1 <= num3 <= 31:
                        fecha = datetime(num1, num2, num3).date()
                        fechas_encontradas.append(fecha)
                        logging.debug(f"Fecha encontrada numérica (YYYY/MM/DD): {fecha} de '{match.group(0)}'")
                    break
                except (ValueError, AttributeError) as e:
                    logging.debug(f"Error parseando fecha numérica: {e}")
                    continue
    
    # ESTRATEGIA 4: Usar dateutil como último recurso (muy permisivo)
    # Buscar patrones que parezcan fechas y probar con dateutil
    palabras_fecha = re.findall(r'\b\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4}\b', texto_limpio)
    for palabra in palabras_fecha:
        try:
            fecha = parse_date(palabra, dayfirst=True).date()
            if 1900 <= fecha.year <= 2100:
                fechas_encontradas.append(fecha)
                logging.debug(f"Fecha encontrada con dateutil: {fecha} de '{palabra}'")
        except:
            continue
    
    # Eliminar duplicados y ordenar
    fechas_unicas = sorted(list(set(fechas_encontradas)))
    
    return fechas_unicas

def parsear_fecha_flexible(fecha_texto: str):
    """
    Intenta parsear texto de fecha con máxima flexibilidad.
    Maneja múltiples formatos y errores de OCR.
    """
    if not fecha_texto:
        return None
    
    # Limpiar el texto
    fecha_texto = limpiar_texto_fecha(fecha_texto.strip())
    
    # Intentar con dateutil primero (más flexible)
    try:
        fecha = parse_date(fecha_texto, dayfirst=True).date()
        if 1900 <= fecha.year <= 2100:
            return fecha
    except:
        pass
    
    # Intentar con formato específico DD/MM/YYYY con varios separadores
    try:
        partes = re.split(r'[/\-\.\s\\]+', fecha_texto)
        if len(partes) == 3:
            dia, mes, año = map(int, partes)
            if año < 100:
                año += 2000 if año < 50 else 1900
            if 1 <= dia <= 31 and 1 <= mes <= 12 and 1900 <= año <= 2100:
                return datetime(año, mes, dia).date()
    except:
        pass
    
    # Intentar formato con nombre de mes
    match = RE_DATE_TEXTO.search(fecha_texto)
    if match:
        try:
            dia = int(match.group(1))
            mes_nombre = match.group(2).lower().strip()
            año = int(match.group(3))
            
            if año < 100:
                año += 2000 if año < 50 else 1900
            
            for key, value in MESES_ESPANOL.items():
                if mes_nombre.startswith(key[:3]) or key.startswith(mes_nombre[:3]):
                    return datetime(año, value, dia).date()
        except:
            pass
    
    return None

def extraer_fecha_mejorada(texto: str, contexto_bancario=False):
    """
    Función mejorada para extraer LA MEJOR fecha de un documento.
    Usa múltiples estrategias y selecciona la más probable.
    """
    fechas = extraer_todas_fechas_texto(texto, contexto_bancario)
    
    if not fechas:
        logging.debug("No se encontraron fechas en el texto")
        return None
    
    # Si encontramos múltiples fechas, aplicar heurísticas para seleccionar la mejor
    if len(fechas) > 1:
        logging.debug(f"Se encontraron {len(fechas)} fechas: {fechas}")
        
        # Filtrar fechas muy antiguas o futuras
        hoy = datetime.now().date()
        fechas_validas = [f for f in fechas if (hoy - timedelta(days=365*5)) <= f <= hoy + timedelta(days=365)]
        
        if not fechas_validas:
            fechas_validas = fechas
        
        # Preferir la fecha más reciente (asumiendo que es la de expedición)
        fecha_seleccionada = max(fechas_validas)
        logging.info(f"Fecha seleccionada (más reciente): {fecha_seleccionada}")
        return fecha_seleccionada
    
    logging.info(f"Fecha encontrada: {fechas[0]}")
    return fechas[0]

def parsear_fecha_texto(fecha_texto: str):
    """Intenta parsear texto de fecha con múltiples formatos (función de compatibilidad)"""
    return parsear_fecha_flexible(fecha_texto)

# ---------------- MEJORADO: Sistema de clasificación de documentos robusto ----------------
DOC_KEYWORDS = {
    "carta": [
        "solicitud", "carta", "solicita devolución", "motivo de la solicitud", 
        "carta de solicitud", "carta de peticion", "representante legal",
        "atentamente", "cordiales saludos", "respetados señores"
    ],
    "rut": [
        "rut", "registro único tributario", "registro unico tributario", 
        "numero de identificacion tributaria", "número de identificación tributaria"
    ],
    "camara_comercio": [
        "cámara de comercio", "camara de comercio", "certificado de existencia",
        "certificado de existencia y representacion", "certificado de representacion legal",
        "matricula mercantil"
    ],
    "cert_bancaria": [
        "certificación bancaria", "certificacion bancaria", "certificado bancario",
        "certificación de cuenta", "certificacion de cuenta", "saldo bancario",
        "certificado de saldo", "entidad financiera", "numero de cuenta",
        "cuenta corriente", "cuenta de ahorros", "banco", "bancaria", "bancario"
    ],
    "recibo_pago": [
        "recibo", "recibos de pago", "planilla", "comprobante de pago",
        "comprobante de transaccion", "voucher de pago", "pago de planilla",
        "transaccion financiera", "numero de recibo"
    ],
    "resolucion": [
        "resolución", "resolucion", "revocó", "revoco", "revocado", "multado",
        "resolucion administrativa", "acto administrativo", "ejecutoriada"
    ],
    "tarjeta_profesional": [
        "tarjeta profesional", "tarjeta prof", "tarjeta del contador",
        "tarjeta del revisor fiscal", "matricula profesional"
    ],
    "acta_consorcial": [
        "acta consorcial", "consorcio", "unión temporal", "union temporal",
        "acta de consorcio", "contrato de union temporal", "joint venture"
    ],
    "contrato": [
        "contrato", "salario integral", "contrato firmado", "contrato de trabajo",
        "contrato laboral", "clausulas contractuales"
    ]
}

def clasificar_documento_robusto(nombre_archivo, texto_ocr):
    """
    Clasificación robusta que combina nombre de archivo y contenido OCR
    con puntuación ponderada y verificación de contexto.
    """
    # Combinar nombre del archivo y texto OCR
    texto_completo = f"{nombre_archivo} {texto_ocr}".lower()
    
    # Normalización extensiva
    texto_completo = texto_completo.replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u')
    texto_completo = re.sub(r'[^\w\s]', ' ', texto_completo)  # Remover puntuación
    texto_completo = re.sub(r'\s+', ' ', texto_completo).strip()
    
    scores = {}
    
    for doc_type, keywords in DOC_KEYWORDS.items():
        score = 0
        for keyword in keywords:
            keyword_normalized = keyword.lower().replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u')
            
            # Puntuación ponderada: más puntos por coincidencias en nombre de archivo
            if keyword_normalized in nombre_archivo.lower():
                score += 3  # Peso alto para nombre de archivo
            if keyword_normalized in texto_completo:
                score += 1  # Peso normal para contenido
            
            # Búsqueda de palabras individuales para contexto
            palabras_keyword = keyword_normalized.split()
            if len(palabras_keyword) > 1:
                # Bonus si todas las palabras de la keyword aparecen (en cualquier orden)
                if all(palabra in texto_completo for palabra in palabras_keyword):
                    score += 2
        
        scores[doc_type] = score
    
    # Obtener el tipo con mayor puntuación
    if not scores or max(scores.values()) == 0:
        return "otro"
    
    tipo_detectado = max(scores, key=scores.get)
    
    # Log para debugging
    if max(scores.values()) > 0:
        logging.debug(f"Clasificación: {tipo_detectado} (puntuación: {scores[tipo_detectado]})")
        logging.debug(f"Todos los scores: {scores}")
    
    return tipo_detectado

def clasificar_texto_keywords(texto: str):
    """Función de compatibilidad - usar solo cuando no hay nombre de archivo"""
    return clasificar_documento_robusto("", texto)

def clasificar_pagina(nombre_archivo, texto: str):
    """Clasificar página con nombre de archivo y contenido"""
    return clasificar_documento_robusto(nombre_archivo, texto)

RE_NIT = re.compile(r"\b(?:NIT[:\s\-]*|NIT\.?[:\s\-]*|NIT\s*)?(\d{6,12})\b")
RE_CEDULA = re.compile(r"\b(?:C[-\s]?C[:\s\-]*|Cédula|Cedula|Cédula de Ciudadanía|CC[:\s\-]*)\s*[:#]?\s*(\d{6,12})\b", re.IGNORECASE)

# ---------------- CHECKLIST ----------------
CHECKLIST = {
    "misional": [
        {
            "id": "carta_representante",
            "titulo": "Carta firmada por el Representate Legal o quien haga sus veces, indicando de forma clara el motivo de la solicitud, relacionando el numero de planillas o tickets, período o fecha de pago, valor pagado y valor a devolver. En casos de apoderados se debe anexar poder debidamente autenticado ante notaría.",
            "keywords": ["carta", "solicitud", "motivo de la solicitud", "planilla", "ticket", "valor a devolver", "apoderado", "poder autenticado"],
            "requerido": True,
            "vigencia_dias": None
        },
        {
            "id": "rut",
            "titulo": "Copia de Rut vigente.",
            "keywords": ["rut", "registro único tributario", "registro unico tributario", "copia rut"],
            "requerido": True,
            "vigencia_dias": None
        },
        {
            "id": "camara_comercio",
            "titulo": "Certificado de Existencia y Representación Legal expedido por la Cámara de Comercio no mayor a 90 días hábiles. Para otro tipo de empleadores no obligados a registrarse en Cámara de Comercio, documento idóneo que acredite la existencia y representación legal.",
            "keywords": ["cámara de comercio", "camara de comercio", "certificado de existencia", "certificado de existencia y representación", "certificado existencia representacion"],
            "requerido": True,
            "vigencia_dias": 90
        },
        {
            "id": "cedula_persona_natural",
            "titulo": "Para empleadores personas naturales copia de la Cédula de Ciudadanía.",
            "keywords": ["cédula", "cedula", "cédula de ciudadanía", "cedula de ciudadania", "cc"],
            "requerido": False,
            "vigencia_dias": None
        },
        {
            "id": "acta_consorcial",
            "titulo": "Para Consorcios y Uniones Temporales se deberá anexar copia del Acta Consorcial firmada por las partes, copia del Rut vigente y copia del Certificado de existencia y Representación Legal de cada uno de los Consorciados o Asociados para personas jurídicas y para personas naturales copia de la cédula de ciudadanía.",
            "keywords": ["acta consorcial", "consorcio", "unión temporal", "union temporal", "acta consorcio", "acta consorcial firmada"],
            "requerido": False,
            "vigencia_dias": None
        },
        {
            "id": "cert_bancaria",
            "titulo": "Certificación Bancaria no mayor a 30 días hábiles de expedición a nombre del peticionario.",
            "keywords": ["certificación bancaria", "certificacion bancaria", "certificado bancario", "certificacion bancaria 30 dias"],
            "requerido": True,
            "vigencia_dias": 30
        },
        {
            "id": "cert_contador",
            "titulo": "Certificación firmada por el Contador o Revisor Fiscal avalando el motivo de la devolución.",
            "keywords": ["certificación firmada", "certificacion firmada", "contador", "revisor fiscal", "certificacion contador"],
            "requerido": True,
            "vigencia_dias": None
        },
        {
            "id": "tarjeta_profesional",
            "titulo": "Copia de la tarjeta profesional del Contador o Revisor Fiscal que firma la certificación.",
            "keywords": ["tarjeta profesional", "tarjeta prof", "tarjeta profesional contador", "tarjeta profesional revisor"],
            "requerido": True,
            "vigencia_dias": None
        },
        {
            "id": "contrato_salario_integral",
            "titulo": "Para devoluciones por error en IBC por Salario Integral anexar copia de contrato firmado por las partes.",
            "keywords": ["contrato", "salario integral", "contrato firmado", "error en IBC"],
            "requerido": False,
            "vigencia_dias": None
        },
        {
            "id": "recibo_pago_fic",
            "titulo": "Recibo de pagos de FIC o Monetización para Contrato de Aprendizaje.",
            "keywords": ["recibo", "fic", "monetizacion", "recibo de pagos", "recibo fic"],
            "requerido": False,
            "vigencia_dias": None
        },
        {
            "id": "resoluciones_multas",
            "titulo": "Copias de las Resoluciones con la cual multaron y revocaron el pago de la Multa debidamente ejecutoriadas a 31 de diciembre de 2019.",
            "keywords": ["resolución", "resolucion", "multado", "revocó", "revoco", "revocaron", "resoluciones multa", "ejecutoriada"],
            "requerido": False,
            "vigencia_dias": None
        }
    ],
    "no_misional": [
        {
            "id": "carta_peticionario",
            "titulo": "Carta de solicitud firmada por el peticionario. En el caso de personas jurídicas deberá ser firmada por el representante legal.",
            "keywords": ["carta", "solicitud", "carta de solicitud", "firmada por el peticionario", "representante legal"],
            "requerido": True,
            "vigencia_dias": None
        },
        {
            "id": "cedula_persona_natural_nm",
            "titulo": "Para personas naturales copia de la cédula de ciudadanía del peticionario.",
            "keywords": ["cédula", "cedula", "cedula de ciudadania", "cédula de ciudadanía", "cc"],
            "requerido": True,
            "vigencia_dias": None
        },
        {
            "id": "camara_comercio_nm",
            "titulo": "Para personas jurídicas, Certificado de Existencia y Representación Legal expedido por la Cámara de Comercio no mayor a 90 días hábiles. Para otro tipo de empleadores no obligados a registrarse en Cámara de Comercio, documento idóneo que acredite la existencia y representación legal.",
            "keywords": ["cámara de comercio", "camara de comercio", "certificado de existencia", "certificado existencia representacion"],
            "requerido": True,
            "vigencia_dias": 90
        },
        {
            "id": "rut_nm",
            "titulo": "Para personas jurídicas, copia de Rut vigente.",
            "keywords": ["rut", "copia rut", "registro único tributario"],
            "requerido": True,
            "vigencia_dias": None
        },
        {
            "id": "cert_bancaria_nm",
            "titulo": "Certificación Bancaria no mayor a 30 días hábiles de expedición a nombre del peticionario.",
            "keywords": ["certificación bancaria", "certificacion bancaria", "certificado bancario"],
            "requerido": True,
            "vigencia_dias": 30
        },
        {
            "id": "recibos_pago_nm",
            "titulo": "Copia del recibo o recibos de pago legibles.",
            "keywords": ["recibo", "recibos de pago", "planilla", "comprobante"],
            "requerido": True,
            "vigencia_dias": None
        },
        {
            "id": "acta_terminacion_convenio",
            "titulo": "Acta de terminación de liquidación del convenio.",
            "keywords": ["acta de terminación", "acta terminacion", "liquidación del convenio", "liquidacion convenio", "terminacion convenio"],
            "requerido": False,
            "vigencia_dias": None
        },
        {
            "id": "factura_produccion_nm",
            "titulo": "Copia de la Factura Electrónica de venta generada por el SENA.",
            "keywords": ["factura", "factura electrónica", "factura electronica", "factura sena", "factura electrónica venta"],
            "requerido": False,
            "vigencia_dias": None
        },
        {
            "id": "cert_autorizacion_coordinador_nm",
            "titulo": "Certificación de autorización de devolución del cupón de pago firmada por el Coordinador del Grupo de Recaudo y Cartera.",
            "keywords": ["autorización de devolución", "autorizacion de devolucion", "coordinador del grupo de recaudo", "coordinador", "certificación de autorización"],
            "requerido": False,
            "vigencia_dias": None
        }
    ]
}

# ---------------- utilidad: normalizar texto (para nombres de archivos) ----------------
STOPWORDS = {"de","del","la","el","los","las","y","s","sa","sas","s.a.s","sas.","sa.","empresa","empresa."}

def quitar_acentos(text: str) -> str:
    """Quita tildes/diacríticos (normalización NFKD)."""
    if not text:
        return ""
    nfkd = unicodedata.normalize("NFKD", text)
    return "".join([c for c in nfkd if not unicodedata.combining(c)])

def normalizar_texto(s: str) -> str:
    """
    Normaliza texto: minúsculas, quita sufijos ruidosos, elimina caracteres no alfanuméricos,
    quita acentos y colapsa espacios.
    """
    if not s:
        return ""
    s = s.lower()

    # quitar extensión .pdf
    s = re.sub(r'\.pdf', '', s, flags=re.IGNORECASE)    

    # patrones ruidosos comunes para truncar el filename (si existen)
    s = re.sub(r'(?i)\b\d{1,2}\-mail\b.*', '', s)      # 01-MAIL...
    s = re.sub(r'(?i)\s*[-\|–]\s*no\..*', '', s)       # - No. ...
    s = re.sub(r'(?i)\s*\bnis\b.*', '', s)             # NIS ...
    s = re.sub(r'(?i)\s*\bradicad[oa]?\b.*', '', s)    # radicado...
    s = re.sub(r'(?i)\b(anexos|respuestas|internas)\b.*', '', s)
    s = re.sub(r'[\(\[\{].*?[\)\]\}]', ' ', s)         # quitar texto entre paréntesis/brackets
    s = re.sub(r'ticketid[_\-]?\d+', ' ', s, flags=re.IGNORECASE)
    s = re.sub(r'\b\d{5,}\b', ' ', s)                 # secuencias largas de números

    # limpieza general de separadores y caracteres
    s = re.sub(r'[_\-\.\,;:\/\\]+', ' ', s)
    s = re.sub(r'[^a-z0-9ñáéíóú\s]', ' ', s)
    s = quitar_acentos(s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def tokens_utiles(s: str):
    s = normalizar_texto(s)
    toks = [t for t in s.split() if t and t not in STOPWORDS]
    return toks

# ---------------- Inferencia directa por substring en filename (DESEADO) ----------------
def inferir_items_desde_nombre(nombre_archivo, tipo_devolucion):
    """
    Inferir items buscándolos directamente como SUBSTRINGs en el filename normalizado.
    - Ignora mayúsculas y tildes.
    - Si encuentra una coincidencia exacta de la keyword (por ejemplo "certificacion bancaria") en el filename,
      retorna ese ItemID.
    - Si múltiples items coinciden, retorna la lista de todos.
    - Si encuentra un marcador de "todo" (anexos, adjuntos, documentos, completo), retorna [] -> evaluar todos.
    """
    if not nombre_archivo:
        return []

    nm = normalizar_texto(nombre_archivo)
    if not nm:
        return []

    # markers que sugieren "todo-en-uno" -> no inferimos individualmente
    todo_markers = {"todo", "completo", "completa", "documentos", "documento", "anexo", "anexos", "adjuntos", "adjunto", "todos"}
    if any(m in nm for m in todo_markers):
        logging.info(f"Filename '{nombre_archivo}' marcado como 'todo' (marker detectado) -> evaluar todos los ítems")
        return []

    checklist_items = CHECKLIST["misional"] if tipo_devolucion == "misional" else CHECKLIST["no_misional"]
    detected = []

    # Recorremos items y sus keywords + título; si alguna frase aparece en nm -> lo asociamos
    for item in checklist_items:
        # construir lista de frases a verificar: título normalizado + keywords normalizadas
        frases = []
        titulo_norm = normalizar_texto(item.get("titulo", ""))
        if titulo_norm:
            frases.append(titulo_norm)
        for kw in item.get("keywords", []):
            kwn = normalizar_texto(kw)
            if kwn:
                frases.append(kwn)

        # deduplicate and sort longer phrases first (priorizar frases compuestas)
        frases = sorted(list(dict.fromkeys(frases)), key=lambda x: -len(x))

        for frase in frases:
            # buscamos la frase como substring en nm
            # ejemplo: frase="certificacion bancaria" -> detecta "certificacion bancaria 123"
            if frase and frase in nm:
                detected.append(item["id"])
                logging.info(f"Archivo '{nombre_archivo}' contiene frase '{frase}' -> item {item['id']}")
                break  # no necesitamos chequear más frases del mismo item

    # si encontramos items -> devolver la lista (sin duplicados)
    if detected:
        return sorted(list(dict.fromkeys(detected)))
    # no se infirió -> devolver [] para indicar "no inferido" (vuelve a evaluar todo)
    return []

# ---------------- NUEVO: Rellenar plantilla Excel con resultados de checklist ----------------
def fill_template_with_checklist(template_path: str,
                                 out_path: str,
                                 df_check: pd.DataFrame,
                                 status_col_letter: str = 'E',
                                 mapping_mode: str = 'auto',  # 'auto'|'itemid'|'title'
                                 fuzzy_threshold: int = 75,
                                 extra_fields: dict = None):
    """
    Rellena la plantilla Excel (hoja activa) con los estados de df_check.
    - template_path: ruta a la plantilla base (.xlsx)
    - out_path: ruta de salida donde se guardará la plantilla rellenada
    - df_check: DataFrame con las columnas al menos: ['ItemID','Item','Estado']
    - status_col_letter: letra de columna donde escribir el estado (por defecto 'E')
    - mapping_mode: 'auto' intenta ItemID y luego título; 'itemid' fuerza ItemID; 'title' fuerza título
    - fuzzy_threshold: umbral (0-100) para fuzzy match si rapidfuzz está instalado
    - extra_fields: dict cell->value para rellenar celdas específicas (ej. {'B2': '01/11/2025', 'B3': 'Empresa X'})
    Returns: dict con conteo escrito y lista de coincidencias no encontradas
    """
    if extra_fields is None:
        extra_fields = {}

    wb = load_workbook(template_path)
    ws = wb.active

    def norm_cell_text(x):
        if x is None:
            return ""
        return re.sub(r'\s+', ' ', str(x).strip()).lower()

    # Build lookup structures from df_check
    # normalize keys
    lookup_by_itemid = {}
    lookup_by_title = {}
    for _, r in df_check.iterrows():
        iid = str(r.get("ItemID") or "").strip()
        title = norm_cell_text(r.get("Item") or "")
        estado = r.get("Estado", "")
        if iid:
            lookup_by_itemid[iid] = estado
        if title:
            lookup_by_title[title] = estado

    written = 0
    not_found = []

    # find header row and possible columns
    first_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    header_map = {}
    for idx, h in enumerate(first_row):
        if h is None:
            continue
        header_map[norm_cell_text(h)] = idx  # 0-based

    # Determine mapping strategy
    use_itemid = False
    use_title = False
    if mapping_mode == 'auto':
        # prefer itemid if template contains a column header 'itemid' (insensitive)
        if any('itemid' in k for k in header_map.keys()):
            use_itemid = True
        else:
            use_title = True
    elif mapping_mode == 'itemid':
        use_itemid = True
    elif mapping_mode == 'title':
        use_title = True

    # If mapping by itemid: find which column contains item ids (case-insensitive)
    itemid_col_idx = None
    title_col_idx = None
    for k, idx in header_map.items():
        if 'itemid' in k:
            itemid_col_idx = idx
        if 'item' == k or 'itemo' in k or 'ítem' in k or 'ítem' in k:
            title_col_idx = idx
        # also try likely labels
        if 'descripcion' in k or 'documento' in k or 'titulo' in k:
            if title_col_idx is None:
                title_col_idx = idx

    status_col_idx = ord(status_col_letter.upper()) - ord('A')  # 0-based

    # If no header mapping for itemid and mapping_mode=itemid, we'll fallback to scanning rows
    # Iterate all rows and try to match
    for row in ws.iter_rows(min_row=2):  # assume header at row 1
        row_idx = row[0].row
        written_this_row = False
        # 1) itemid mapping
        if use_itemid:
            if itemid_col_idx is not None and itemid_col_idx < len(row):
                cell_val = row[itemid_col_idx].value
                if cell_val is not None:
                    key = str(cell_val).strip()
                    if key in lookup_by_itemid:
                        ws.cell(row=row_idx, column=status_col_idx+1, value=lookup_by_itemid[key])
                        written += 1
                        written_this_row = True
                    else:
                        # fuzzy attempt on itemid strings (if available)
                        if HAVE_RAPIDFUZZ:
                            best = None
                            best_score = 0
                            for iid_k, est in lookup_by_itemid.items():
                                score = fuzz.partial_ratio(str(iid_k), str(key))
                                if score > best_score:
                                    best_score = score
                                    best = iid_k
                            if best_score >= fuzzy_threshold:
                                ws.cell(row=row_idx, column=status_col_idx+1, value=lookup_by_itemid[best])
                                written += 1
                                written_this_row = True
            else:
                # fallback: scan entire row values for itemid text match
                for c in row:
                    val = c.value
                    if val is None:
                        continue
                    val_s = str(val).strip()
                    if val_s in lookup_by_itemid:
                        ws.cell(row=row_idx, column=status_col_idx+1, value=lookup_by_itemid[val_s])
                        written += 1
                        written_this_row = True
                        break
        # 2) title mapping
        if (not written_this_row) and use_title:
            # Prefer title col if known
            cell_text = ""
            if title_col_idx is not None and title_col_idx < len(row):
                cell_text = norm_cell_text(row[title_col_idx].value)
            else:
                # fallback to scanning a few likely columns (0..5)
                possible = []
                for c in row[:6]:
                    possible.append(norm_cell_text(c.value))
                cell_text = " ".join(possible)
            matched = False
            # exact / substring match first
            for title_norm, estado in lookup_by_title.items():
                if title_norm and title_norm in cell_text:
                    ws.cell(row=row_idx, column=status_col_idx+1, value=estado)
                    written += 1
                    matched = True
                    break
            if not matched and HAVE_RAPIDFUZZ:
                # fuzzy best match (compare title_norm vs cell_text)
                best_title = None
                best_score = 0
                for title_norm, estado in lookup_by_title.items():
                    try:
                        score = fuzz.partial_ratio(title_norm, cell_text)
                    except Exception:
                        score = 0
                    if score > best_score:
                        best_score = score
                        best_title = title_norm
                if best_score >= fuzzy_threshold:
                    ws.cell(row=row_idx, column=status_col_idx+1, value=lookup_by_title[best_title])
                    written += 1
                    matched = True
            if not matched:
                # nothing matched -> collect for report
                not_found.append((row_idx, cell_text))

    # write extra fields (cell addresses)
    for cell_addr, val in (extra_fields or {}).items():
        try:
            ws[cell_addr] = val
        except Exception as e:
            logging.warning(f"Error escribiendo campo extra {cell_addr}: {e}")

    # save
    wb.save(out_path)
    return {"written": written, "not_found": not_found}

# ---------------- procesamiento PDF -> páginas ----------------
def procesar_pdf_a_paginas(path_pdf, poppler_path=None):
    try:
        images = convert_from_path(path_pdf, dpi=200, poppler_path=poppler_path) if poppler_path else convert_from_path(path_pdf, dpi=200)
        return images
    except Exception as e:
        logging.error(f"Error al convertir PDF a imágenes ({path_pdf}): {e}")
        raise

# ---------------- MEJORADO: Extracción de información con clasificación robusta ----------------
def extraer_info_por_pagina(path_pdf, fecha_recepcion_dt, poppler_path=None):
    images = procesar_pdf_a_paginas(path_pdf, poppler_path=poppler_path)
    resultados = []
    nombre_archivo = os.path.basename(path_pdf)
    
    for i, img in enumerate(images, start=1):
        texto, conf = ocr_image_and_confidence(img)
        
        # USAR CLASIFICACIÓN ROBUSTA con nombre de archivo y contenido
        tipo = clasificar_pagina(nombre_archivo, texto)
        
        nit = None
        m = RE_NIT.search(texto)
        if m:
            nit = m.group(1)
        else:
            m2 = RE_CEDULA.search(texto)
            if m2:
                nit = m2.group(1)
        
        # MEJORADO: Detección de fecha con logging detallado
        fecha_doc = None
        # Detectar si es documento bancario
        contexto_bancario = any(kw in texto.lower() for kw in 
                               ["certificacion bancaria", "certificado bancario", "banco", 
                                "cuenta", "bancaria", "entidad financiera", "certificación bancaria"])
        
        if contexto_bancario:
            logging.info(f"Página {i} de {nombre_archivo}: Detectado contexto bancario")
        
        # Extraer fecha con estrategia mejorada
        fecha_doc = extraer_fecha_mejorada(texto, contexto_bancario=contexto_bancario)
        
        if fecha_doc:
            logging.info(f"Página {i} de {nombre_archivo}: Fecha extraída: {fecha_doc}")
        else:
            logging.warning(f"Página {i} de {nombre_archivo}: No se pudo extraer fecha")
            # Mostrar fragmento del texto para debug (primeros 300 caracteres)
            if texto:
                logging.debug(f"Fragmento de texto: {texto[:300]}")
        
        firma = detectar_firma_manuscrita(img)
        resultados.append({
            "archivo": nombre_archivo,
            "ruta_archivo": path_pdf,
            "pagina": i,
            "imagen": img,
            "texto": texto,
            "ocr_conf": conf,
            "tipo_detectado": tipo,
            "nit_o_cedula": nit,
            "fecha_documento": fecha_doc.isoformat() if fecha_doc else None,
            "firma_manuscrita": firma
        })
    return resultados

# ---------------- Agrupar páginas contiguas del mismo 'tipo_detectado' ----------------
def agrupar_paginas_en_documentos(paginas_info):
    if not paginas_info:
        return []
    documentos = []
    current = None
    for p in paginas_info:
        tipo = p.get("tipo_detectado") or "otro"
        if current is None:
            current = {
                "archivo": p["archivo"],
                "ruta_archivo": p["ruta_archivo"],
                "paginas": [p["pagina"]],
                "texto": p["texto"] or "",
                "ocr_confs": [p.get("ocr_conf") or 0.0],
                "tipo_detectado": tipo,
                "nit_o_cedula": p.get("nit_o_cedula"),
                "fecha_documento": p.get("fecha_documento"),
                "firma_manuscrita": bool(p.get("firma_manuscrita")),
            }
            continue
        if tipo == current["tipo_detectado"]:
            current["paginas"].append(p["pagina"])
            current["texto"] += "\n\n" + (p.get("texto") or "")
            current["ocr_confs"].append(p.get("ocr_conf") or 0.0)
            if not current["nit_o_cedula"] and p.get("nit_o_cedula"):
                current["nit_o_cedula"] = p.get("nit_o_cedula")
            if not current["fecha_documento"] and p.get("fecha_documento"):
                current["fecha_documento"] = p.get("fecha_documento")
            current["firma_manuscrita"] = current["firma_manuscrita"] or bool(p.get("firma_manuscrita"))
        else:
            avg_conf = float(np.mean(current["ocr_confs"])) if current["ocr_confs"] else 0.0
            documentos.append({
                "archivo": current["archivo"],
                "ruta_archivo": current["ruta_archivo"],
                "paginas": current["paginas"],
                "texto": current["texto"],
                "ocr_conf": avg_conf,
                "tipo_detectado": current["tipo_detectado"],
                "nit_o_cedula": current["nit_o_cedula"],
                "fecha_documento": current["fecha_documento"],
                "firma_manuscrita": current["firma_manuscrita"]
            })
            current = {
                "archivo": p["archivo"],
                "ruta_archivo": p["ruta_archivo"],
                "paginas": [p["pagina"]],
                "texto": p["texto"] or "",
                "ocr_confs": [p.get("ocr_conf") or 0.0],
                "tipo_detectado": tipo,
                "nit_o_cedula": p.get("nit_o_cedula"),
                "fecha_documento": p.get("fecha_documento"),
                "firma_manuscrita": bool(p.get("firma_manuscrita")),
            }
    if current:
        avg_conf = float(np.mean(current["ocr_confs"])) if current["ocr_confs"] else 0.0
        documentos.append({
            "archivo": current["archivo"],
            "ruta_archivo": current["ruta_archivo"],
            "paginas": current["paginas"],
            "texto": current["texto"],
            "ocr_conf": avg_conf,
            "tipo_detectado": current["tipo_detectado"],
            "nit_o_cedula": current["nit_o_cedula"],
            "fecha_documento": current["fecha_documento"],
            "firma_manuscrita": current["firma_manuscrita"]
        })
    return documentos

# ---------------- classifier (compatibilidad) ----------------
def clasificar_texto_keywords_old(texto: str):
    """Función antigua mantenida por compatibilidad"""
    texto_low = (texto or "").lower()
    scores = {}
    for doc_type, keys in DOC_KEYWORDS.items():
        s = sum(1 for k in keys if k in texto_low)
        scores[doc_type] = s
    best = max(scores.items(), key=lambda x: x[1])
    return best[0] if best[1] > 0 else "otro"

# ---------------- CORREGIDO: Evaluar item con manejo completo de tipos de fecha ----------------
# ---------------- CORREGIDO: Evaluar item con manejo completo de tipos de fecha ----------------
def evaluar_item(item, documentos_detectados, fecha_recepcion_dt, peticionario_tipo="persona_juridica", archivo_inferido=None):
    archivos = []
    observaciones = []
    estado = "Falta"

    # ---------------- MODIFICACIÓN SOLICITADA ----------------
    # Para no misional: si es persona natural, no aplicar items de persona jurídica
    if peticionario_tipo == "persona_natural":
        if item.get("id") in ["camara_comercio_nm", "rut_nm"]:
            return "N/A", None, "No aplica para persona natural"
    # ---------------- FIN MODIFICACIÓN ----------------
    
    if item.get("id") in ["cedula_persona_natural","cedula_persona_natural_nm"] and peticionario_tipo != "persona_natural":
        return "N/A", None, "Aplica solo si peticionario es persona natural"
    if item.get("id") == "acta_consorcial" and peticionario_tipo != "consorcio":
        return "N/A", None, "Aplica solo si peticionario es consorcio/unión temporal"

    # CORRECCIÓN: Función auxiliar simplificada para manejar tipos de fecha
    def obtener_fecha_date(fecha_obj):
        """Convierte cualquier tipo de fecha a datetime.date de manera segura"""
        if fecha_obj is None:
            return None
        
        # Si ya es date, devolver directamente
        if type(fecha_obj).__name__ == 'date':
            return fecha_obj
        
        # Si es datetime, extraer la parte date
        if type(fecha_obj).__name__ == 'datetime':
            return fecha_obj.date()
        
        # Si es string, intentar parsear
        if isinstance(fecha_obj, str):
            try:
                parsed = parse_date(fecha_obj, dayfirst=True)
                return parsed.date() if hasattr(parsed, 'date') else parsed
            except:
                return None
        
        return None

    for doc in documentos_detectados:
        texto = (doc.get("texto") or "").lower()
        score = 0
        for kw in item.get("keywords", []):
            if kw.lower() in texto:
                score += 1
        
        # Si el ítem fue inferido por nombre de archivo, ser más permisivo
        es_inferido = archivo_inferido and item["id"] in archivo_inferido
        
        # Si fue inferido, requerir menos coincidencias
        if (es_inferido and score >= 1) or (not es_inferido and score >= 2):
            paginas_str = ",".join(str(p) for p in doc.get("paginas", []))
            archivos.append(f"{doc['archivo']} (p{paginas_str})")
            
            if es_inferido:
                observaciones.append("Documento asociado por inferencia de nombre de archivo")
            
            vig = item.get("vigencia_dias")
            if vig:
                fecha_doc = None
                
                # Intentar obtener la fecha del documento de manera robusta
                if doc.get("fecha_documento"):
                    fecha_doc = obtener_fecha_date(doc.get("fecha_documento"))
                
                # Si no hay fecha, intentar extraerla del texto
                if not fecha_doc:
                    contexto_bancario = "bancaria" in item.get("id", "") or "cert_bancaria" in item.get("id", "")
                    texto_completo = doc.get("texto", "")
                    fecha_redetectada = extraer_fecha_mejorada(texto_completo, contexto_bancario)
                    
                    if fecha_redetectada:
                        fecha_doc = fecha_redetectada
                        doc["fecha_documento"] = fecha_doc.isoformat()
                        observaciones.append("Fecha redetectada con método mejorado")
                
                # Validar vigencia si tenemos fecha
                if fecha_doc:
                    try:
                        # CORRECCIÓN: Manejo consistente de todas las fechas
                        fecha_recepcion_date = obtener_fecha_date(fecha_recepcion_dt)
                        
                        if fecha_recepcion_date is None:
                            estado = "Revisión"
                            observaciones.append("No se pudo determinar la fecha de recepción")
                        else:
                            # Convertir fecha_doc a datetime para fecha_limite_habiles
                            if hasattr(fecha_doc, 'date'):
                                fecha_doc_datetime = datetime.combine(fecha_doc, datetime.min.time())
                            else:
                                fecha_doc_datetime = datetime.combine(fecha_doc, datetime.min.time())
                            
                            # FECHA LÍMITE = fecha de expedición + días hábiles
                            fecha_limite_validez = fecha_limite_habiles(fecha_doc_datetime, vig).date()
                            
                            # LÓGICA CORREGIDA: La fecha de recepción debe ser <= fecha límite
                            if fecha_recepcion_date <= fecha_limite_validez:
                                estado = "C"
                                observaciones.append(f"Vigente: Recepción {fecha_recepcion_date} ≤ Límite {fecha_limite_validez} (Expedición {fecha_doc} + {vig}d hábiles)")
                            else:
                                estado = "Falta"
                                observaciones.append(f"Vencido: Recepción {fecha_recepcion_date} > Límite {fecha_limite_validez} (Expedición {fecha_doc} + {vig}d hábiles)")
                    except Exception as e:
                        estado = "Revisión"
                        observaciones.append(f"Error validando vigencia: {str(e)}")
                        logging.error(f"Error detallado en validación de vigencia: {e}")
                else:
                    estado = "Revisión"
                    observaciones.append("Fecha no encontrada para validar vigencia")
            else:
                estado = "C"

            conf = doc.get("ocr_conf", 0.0)
            if conf < 30 or len((doc.get("texto") or "").strip()) < 40:
                estado = "Revisión"
                observaciones.append(f"OCR baja/confianza={conf:.1f}")

            if item["id"] in ("carta_representante","carta_peticionario","cert_contador","cert_contador"):
                if not doc.get("firma_manuscrita"):
                    observaciones.append("Firma manuscrita no detectada; verificar firma digital o firma escaneada")
                    if estado == "C":
                        estado = "Revisión"

    if not archivos:
        if not item.get("requerido", True):
            estado = "N/A"
        else:
            estado = "Falta"
    return estado, archivos if archivos else None, "; ".join(observaciones) if observaciones else None

# ---------------- generar_checklist (con allowed_item_ids) ----------------
def generar_checklist(documentos_detectados, tipo_devolucion, fecha_recepcion_str, peticionario_tipo="persona_juridica", allowed_item_ids=None):
    # CORRECCIÓN: Manejo robusto de la fecha de recepción
    try:
        if isinstance(fecha_recepcion_str, (datetime, datetime.date)):
            fecha_recepcion_dt = fecha_recepcion_str
        else:
            fecha_recepcion_dt = parse_date(fecha_recepcion_str, dayfirst=True)
    except Exception as e:
        logging.error(f"Error parseando fecha de recepción '{fecha_recepcion_str}': {e}")
        # Usar fecha actual como fallback
        fecha_recepcion_dt = datetime.now()
    
    todos_items = CHECKLIST["misional"] if tipo_devolucion == "misional" else CHECKLIST["no_misional"]

    if allowed_item_ids:
        items = [item for item in todos_items if item["id"] in set(allowed_item_ids)]
    else:
        items = todos_items

    filas = []
    for item in items:
        estado, archivos, obs = evaluar_item(item, documentos_detectados, fecha_recepcion_dt, 
                                            peticionario_tipo=peticionario_tipo,
                                            archivo_inferido=allowed_item_ids)
        filas.append({
            "TipoDevolucion": tipo_devolucion,
            "PeticionarioTipo": peticionario_tipo,
            "ItemID": item["id"],
            "Item": item["titulo"],
            "Requerido": item.get("requerido", True),
            "Estado": estado,
            "ArchivosFuente": ", ".join(archivos) if archivos else None,
            "Observaciones": obs
        })
    
    return pd.DataFrame(filas)

# ---------------- procesamiento batch de carpeta (CLI) ----------------
def procesar_carpeta(folder_path, fecha_recepcion_str, tipo_devolucion, peticionario_tipo="persona_juridica", poppler_path=None, tesseract_cmd=None, out_excel="informe_devoluciones.xlsx", template_misional=TEMPLATE_MISIONAL, template_no_misional=TEMPLATE_NO_MISIONAL):
    if tesseract_cmd:
        pytesseract.pytesseract.tesseract_cmd = tesseract_cmd

    todas_las_filas = []
    resumen_archivos = []
    pdf_files = [f for f in os.listdir(folder_path) if f.lower().endswith(".pdf")]
    if not pdf_files:
        logging.info("No hay PDFs en la carpeta indicada.")
        return None

    for fname in pdf_files:
        full_path = os.path.join(folder_path, fname)
        logging.info(f"Procesando: {fname} ...")
        try:
            paginas_info = extraer_info_por_pagina(full_path, parse_date(fecha_recepcion_str, dayfirst=True), poppler_path=poppler_path)
            documentos = agrupar_paginas_en_documentos(paginas_info)

            # inferir items desde nombre (ahora: búsqueda directa de frases/substring)
            allowed_ids = inferir_items_desde_nombre(fname, tipo_devolucion)
            if allowed_ids:
                logging.info(f"Inferencia desde nombre: {fname} -> items {allowed_ids}")
                df_check = generar_checklist(documentos, tipo_devolucion, fecha_recepcion_str, peticionario_tipo, allowed_item_ids=allowed_ids)
            else:
                logging.info(f"No se infirió item específico para {fname} -> evaluando todos los ítems")
                df_check = generar_checklist(documentos, tipo_devolucion, fecha_recepcion_str, peticionario_tipo)

            df_check["SolicitudArchivo"] = fname
            df_check["NombreArchivo"] = fname
            df_check["InferredItemIDs"] = ",".join(allowed_ids) if allowed_ids else None

            todas_las_filas.append(df_check)
            resumen_archivos.append({
                "archivo": fname,
                "paginas_detectadas": len(paginas_info),
                "documentos_detectados": len(documentos),
                "hash_sha256": sha256_file(full_path)
            })
        except Exception as e:
            logging.exception(f"ERROR procesando {fname}: {e}")
            todas_las_filas.append(pd.DataFrame([{
                "TipoDevolucion": tipo_devolucion,
                "PeticionarioTipo": peticionario_tipo,
                "ItemID": "ERROR",
                "Item": f"ERROR al procesar {fname}",
                "Requerido": True,
                "Estado": "ERROR",
                "ArchivosFuente": fname,
                "Observaciones": str(e),
                "SolicitudArchivo": fname,
                "NombreArchivo": fname,
                "InferredItemIDs": None
            }]))

    if todas_las_filas:
        df_total = pd.concat(todas_las_filas, ignore_index=True)
    else:
        df_total = pd.DataFrame()
    df_resumen = pd.DataFrame(resumen_archivos)

    # Guardar auditoría como antes
    with pd.ExcelWriter(out_excel, engine="openpyxl") as writer:
        df_total.to_excel(writer, sheet_name="Checklist", index=False)
        df_resumen.to_excel(writer, sheet_name="ResumenArchivos", index=False)
    logging.info(f"Proceso terminado. Informe generado: {out_excel}")

    # --- RELLENAR PLANTILLA (si existe) ---
    try:
        plantilla = template_misional if tipo_devolucion == "misional" else template_no_misional
        if os.path.exists(plantilla):
            # Construir extra_fields por defecto
            extra_fields = {
                # ejemplo: poner fecha recepción en B2, tipo en B3, cantidad de archivos en B4
                # Ajusta según tu plantilla real
                "B2": fecha_recepcion_str,
                "B3": tipo_devolucion,
                "B4": len(pdf_files),
                "B5": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            out_filled = os.path.join(os.path.dirname(out_excel) or ".", f"informe_relleno_{tipo_devolucion}.xlsx")
            # Llamada con auto mapping y fuzzy fallback
            res = fill_template_with_checklist(plantilla, out_filled, df_total, status_col_letter='E', mapping_mode='auto', fuzzy_threshold=75, extra_fields=extra_fields)
            logging.info(f"Plantilla rellenada y guardada en {out_filled} (escritas={res.get('written')})")
        else:
            logging.info(f"No se encontró plantilla {plantilla} — se omitió relleno de plantilla.")
    except Exception as e:
        logging.exception(f"Error al rellenar plantilla: {e}")

    return out_excel

# ---------------- Streamlit UI (con agrupamiento e inferencia robusta) ----------------
def run_streamlit_app(default_poppler=POPPLER_PATH, default_tesseract=TESSERACT_CMD):
    if not STREAMLIT_AVAILABLE:
        print("Streamlit no está instalado. Instala streamlit para usar la interfaz: pip install streamlit")
        return

    st.set_page_config(page_title="Procesador Devoluciones SENA", layout="wide")
    st.title("Procesador Automático — Lista de Chequeo Devoluciones SENA (UI)")
    st.markdown("Sube PDFs (uno o varios). Cada PDF se interpreta como una solicitud. Revisa OCR y descarga informe final.")

    with st.sidebar:
        st.header("Parámetros")
        fecha_rec = st.date_input("Fecha de recepción", value=datetime.today())
        tipo_dev = st.selectbox("Tipo de devolución", options=["misional","no_misional"])
        pet_tipo = st.selectbox("Tipo de peticionario", options=["persona_juridica","persona_natural","consorcio"])
        poppler_in = st.text_input("Ruta a poppler (si Windows)", value=default_poppler or "")
        tesseract_in = st.text_input("Ruta Tesseract (si necesario)", value=default_tesseract or "")

    st.info("Arrastra y suelta uno o varios archivos PDF. Cada PDF = 1 solicitud.")
    uploaded = st.file_uploader("Sube archivos PDF", type="pdf", accept_multiple_files=True)

    if uploaded:
        st.write(f"{len(uploaded)} archivo(s) cargado(s).")
        resultados_check_total = []
        resumen_archivos = []
        for uf in uploaded:
            st.write("---")
            st.subheader(f"Archivo: {uf.name}")
            try:
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
                tmp.write(uf.read())
                tmp.flush()
                tmp.close()
                tmp_path = tmp.name

                images = convert_from_path(tmp_path, dpi=200, poppler_path=poppler_in or None) if (poppler_in or default_poppler) else convert_from_path(tmp_path, dpi=200)
                st.write(f"Páginas detectadas: {len(images)}")

                if st.checkbox(f"Ver vista previa OCR de {uf.name}", key=f"vp_{uf.name}"):
                    col1, col2 = st.columns(2)
                    for i, img in enumerate(images[:6], start=1):
                        texto, conf = ocr_image_and_confidence(img)
                        with col1:
                            st.image(img, caption=f"{uf.name} - p{i}", use_column_width=True)
                        with col2:
                            st.text_area(f"Texto p{i} (conf={conf:.1f})", value=texto[:2000], height=200)

                paginas_info = []
                for i, img in enumerate(images, start=1):
                    texto, conf = ocr_image_and_confidence(img)
                    # USAR CLASIFICACIÓN ROBUSTA en Streamlit
                    tipo = clasificar_pagina(uf.name, texto)
                    nit = None
                    m = RE_NIT.search(texto)
                    if m:
                        nit = m.group(1)
                    else:
                        m2 = RE_CEDULA.search(texto)
                        if m2:
                            nit = m2.group(1)
                    
                    # MEJORADO: Detección de fecha en Streamlit
                    fecha_doc = None
                    contexto_bancario = any(kw in texto.lower() for kw in ["certificacion bancaria", "certificado bancario", "banco", "cuenta", "bancaria", "entidad financiera"])
                    fecha_doc = extraer_fecha_mejorada(texto, contexto_bancario=contexto_bancario)
                    
                    if not fecha_doc:
                        md = RE_DATE.search(texto)
                        if md:
                            try:
                                for group_num in range(1, 12):
                                    if md.group(group_num):
                                        try:
                                            fecha_doc = parse_date(md.group(group_num), dayfirst=True).date()
                                            break
                                        except:
                                            continue
                            except:
                                fecha_doc = None
                    
                    firma = detectar_firma_manuscrita(img)
                    paginas_info.append({
                        "archivo": uf.name,
                        "ruta_archivo": tmp_path,
                        "pagina": i,
                        "imagen": img,
                        "texto": texto,
                        "ocr_conf": conf,
                        "tipo_detectado": tipo,
                        "nit_o_cedula": nit,
                        "fecha_documento": fecha_doc.isoformat() if fecha_doc else None,
                        "firma_manuscrita": firma
                    })

                documentos = agrupar_paginas_en_documentos(paginas_info)
                st.write(f"Documentos detectados (grupos de páginas contiguas): {len(documentos)}")
                doc_preview = []
                for d in documentos:
                    doc_preview.append({
                        "tipo": d["tipo_detectado"],
                        "paginas": "-".join(str(p) for p in d["paginas"]),
                        "conf_promedio": round(d.get("ocr_conf",0.0),1),
                        "nit": d.get("nit_o_cedula"),
                        "fecha": d.get("fecha_documento"),
                        "firma": d.get("firma_manuscrita")
                    })
                st.dataframe(pd.DataFrame(doc_preview), height=200)

                # inferencia: ahora por substring exacto en filename (normalizado)
                allowed_ids = inferir_items_desde_nombre(uf.name, tipo_dev)
                if allowed_ids:
                    st.info(f"Inferred items from filename: {allowed_ids}")
                    df_check = generar_checklist(documentos, tipo_dev, fecha_rec, peticionario_tipo=pet_tipo, allowed_item_ids=allowed_ids)
                else:
                    st.info("No se infirió item desde el nombre del archivo. Evaluando todos los ítems.")
                    df_check = generar_checklist(documentos, tipo_dev, fecha_rec.strftime("%d/%m/%Y"), peticionario_tipo=pet_tipo)

                st.write("Resultados (por ítem):")
                st.dataframe(df_check, height=300)

                df_check["SolicitudArchivo"] = uf.name
                df_check["NombreArchivo"] = uf.name
                df_check["InferredItemIDs"] = ",".join(allowed_ids) if allowed_ids else None

                resultados_check_total.append(df_check)
                resumen_archivos.append({"archivo": uf.name, "paginas": len(images), "documentos_detectados": len(documentos), "sha256": sha256_file(tmp_path)})
                try:
                    os.remove(tmp_path)
                except:
                    pass

            except Exception as e:
                st.error(f"Error procesando {uf.name}: {e}")
                logging.exception(f"Error procesando {uf.name}: {e}")

        if resultados_check_total:
            df_total = pd.concat(resultados_check_total, ignore_index=True)
            df_resumen = pd.DataFrame(resumen_archivos)
            towrite = io.BytesIO()
            with pd.ExcelWriter(towrite, engine="openpyxl") as writer:
                df_total.to_excel(writer, sheet_name="Checklist", index=False)
                df_resumen.to_excel(writer, sheet_name="ResumenArchivos", index=False)
            towrite.seek(0)
            st.download_button("Descargar informe Excel", data=towrite, file_name="informe_devoluciones.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            st.success("Informe listo.")
    else:
        st.write("No se han subido archivos aún.")

    st.markdown("---")
    st.caption("Versión mejorada con detección robusta de fechas y documentos en múltiples formatos.")

# ---------------- Entrypoint: CLI o UI según invocación ----------------
def main_cli():
    parser = argparse.ArgumentParser(description="Procesador devoluciones SENA - MVP (Versión mejorada)")
    parser.add_argument("carpeta_pdfs", help="Carpeta que contiene los PDFs (cada PDF = 1 solicitud)")
    parser.add_argument("fecha_recepcion", help="Fecha de recepción (dd/mm/yyyy)")
    parser.add_argument("tipo_devolucion", choices=["misional","no_misional"], help="Tipo de devolución")
    parser.add_argument("--peticionario", choices=["persona_natural","persona_juridica","consorcio"], default="persona_juridica", help="Tipo de peticionario (afecta N/A)")
    parser.add_argument("--poppler", default=None, help="Ruta a bin de poppler (si es necesario en Windows)")
    parser.add_argument("--tesseract", default=None, help="Ruta al ejecutable de tesseract (si no está en PATH)")
    parser.add_argument("--out", default="informe_devoluciones.xlsx", help="Archivo Excel de salida")
    parser.add_argument("--template_misional", default=TEMPLATE_MISIONAL, help="Plantilla Excel misional (opcional)")
    parser.add_argument("--template_no_misional", default=TEMPLATE_NO_MISIONAL, help="Plantilla Excel no_misional (opcional)")
    parser.add_argument("--debug", action="store_true", help="Activar modo debug con logging detallado")
    args = parser.parse_args()

    # Activar logging detallado si se solicita
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
        logging.info("Modo DEBUG activado - se mostrarán todos los detalles de detección de fechas")

    poppler_path = args.poppler or POPPLER_PATH
    tesseract_cmd = args.tesseract or TESSERACT_CMD

    procesar_carpeta(args.carpeta_pdfs, args.fecha_recepcion, args.tipo_devolucion,
                     peticionario_tipo=args.peticionario, poppler_path=poppler_path,
                     tesseract_cmd=tesseract_cmd, out_excel=args.out,
                     template_misional=args.template_misional,
                     template_no_misional=args.template_no_misional)

if __name__ == "__main__":
    # Si streamlit está disponible y el script se invoca sin args preferimos UI para compatibilidad
    if STREAMLIT_AVAILABLE and (len(sys.argv) == 1 or "streamlit" in sys.argv[0].lower()):
        run_streamlit_app(default_poppler=POPPLER_PATH, default_tesseract=TESSERACT_CMD)
    else:
        main_cli()